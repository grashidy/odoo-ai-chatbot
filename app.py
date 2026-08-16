import xmlrpc.client, json, os, time, re, logging, threading, io, base64, socket, datetime
from pathlib import Path
from flask import Flask, render_template, request, jsonify, Response, stream_with_context
from openai import OpenAI
try:
    from groq import Groq as _GroqClient  # kept only for Whisper voice transcription
except ImportError:
    _GroqClient = None
try:
    import requests as _requests
except ImportError:
    _requests = None

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")

# Global socket timeout prevents xmlrpc calls from hanging indefinitely
socket.setdefaulttimeout(55)

# ── AI Provider Manager ────────────────────────────────────────────────────────
# Models are fully configurable via Railway environment variables.
# Never hard-code model names — set GEMINI_MODEL / GROQ_MODEL in Railway Variables.
GEMINI_BASE_URL = "https://generativelanguage.googleapis.com/v1beta/openai/"
GROQ_BASE_URL   = "https://api.groq.com/openai/v1"

class AIProviderManager:
    """
    Centralized AI provider with automatic Gemini → Groq fallback.

    Decision table:
      401 / 403  →  auth/config error        →  fallback immediately
      404        →  model unavailable         →  fallback immediately
      429        →  rate limit / daily quota  →  fallback immediately
      5xx        →  provider server error     →  fallback immediately
      timeout    →  request timed out         →  fallback immediately
      400        →  bad request (code bug)    →  NO fallback, show error
    """

    _USER_MESSAGES = {
        "auth_error":        "⚠️ Invalid API key. Go to Railway → Variables and check OPENAI_API_KEY / GEMINI_API_KEY / GROQ_API_KEY are correct and not revoked.",
        "model_unavailable": "⚠️ The requested AI model is unavailable. Trying backup provider…",
        "rate_limit":        "⚠️ AI rate limit reached. Please wait a moment and try again.",
        "daily_quota":       "⚠️ All AI providers are unavailable (quota exhausted or invalid keys). Fix: go to Railway → Variables and update GEMINI_API_KEY with a fresh key from aistudio.google.com (free). Then check OPENAI_API_KEY is correct.",
        "server_error":      "⚠️ AI service temporarily unavailable. Please try again shortly.",
        "timeout":           "⚠️ AI request timed out. Please try again.",
        "bad_request":       "⚠️ The AI received an invalid request. Please rephrase your question.",
        "no_providers":      "⚠️ No AI providers configured. Set OPENAI_API_KEY, GEMINI_API_KEY, or GROQ_API_KEY in Railway Variables.",
        "unknown_error":     "⚠️ AI service temporarily unavailable. Please try again shortly.",
    }

    def __init__(self):
        self.openai   = self._build_openai()   # primary
        self.primary  = self._build_gemini()   # secondary
        self.fallback = self._build_groq()     # tertiary
        self._log_startup()

    # ── Provider builders ────────────────────────────────────────────────────

    def _build_openai(self):
        key = os.environ.get("OPENAI_API_KEY", "").strip()
        if not key:
            return None
        model = os.environ.get("OPENAI_MODEL", "gpt-4o")
        return {"name": "OpenAI", "model": model,
                "client": OpenAI(api_key=key)}   # default base_url = api.openai.com

    def _build_gemini(self):
        key = os.environ.get("GEMINI_API_KEY", "").strip()
        if not key:
            return None
        model = os.environ.get("GEMINI_MODEL", "gemini-2.0-flash-lite")
        return {"name": "Gemini", "model": model,
                "client": OpenAI(api_key=key, base_url=GEMINI_BASE_URL)}

    def _build_groq(self):
        _kf = Path(__file__).parent / ".groq_key"
        key = (os.environ.get("GROQ_API_KEY", "").strip()
               or (_kf.read_text(encoding="utf-8").strip() if _kf.exists() else ""))
        if not key:
            return None
        model = os.environ.get("GROQ_MODEL", "llama-3.1-8b-instant")
        return {"name": "Groq", "model": model,
                "client": OpenAI(api_key=key, base_url=GROQ_BASE_URL)}

    # ── Error classification ─────────────────────────────────────────────────

    def classify_error(self, exc):
        """Returns (category: str, should_fallback: bool)."""
        msg = str(exc)
        if "401" in msg or "403" in msg:
            return "auth_error", True
        if "404" in msg:
            return "model_unavailable", True
        if "429" in msg or "rate_limit" in msg.lower():
            if any(x in msg.lower() for x in ("per day", "daily", "exceeded", "tpd")):
                return "daily_quota", True
            return "rate_limit", True
        if any(f"{c}" in msg for c in (500, 502, 503, 504)):
            return "server_error", True
        if "timeout" in msg.lower() or "timed out" in msg.lower():
            return "timeout", True
        if "400" in msg:
            return "bad_request", True        # try next provider — tools schema may differ
        return "unknown_error", True

    def user_message(self, category):
        return self._USER_MESSAGES.get(category, self._USER_MESSAGES["unknown_error"])

    # ── Provider iteration ───────────────────────────────────────────────────

    def providers(self, override_key=None):
        """Return providers in priority order: OpenAI → Gemini → Groq."""
        plist = []
        if override_key:
            plist.append({"name": "Custom", "model": "llama-3.1-8b-instant",
                          "client": OpenAI(api_key=override_key, base_url=GROQ_BASE_URL)})
        else:
            if self.openai:   plist.append(self.openai)
            if self.primary:  plist.append(self.primary)
            if self.fallback: plist.append(self.fallback)
        return plist

    def best_client(self, override_key=None):
        """Return (client, model) of the first available provider — for simple one-shot calls."""
        for p in self.providers(override_key):
            return p["client"], p["model"]
        return None, None

    # ── Startup log ─────────────────────────────────────────────────────────

    def _log_startup(self):
        logging.warning("══ AI PROVIDER STATUS ══════════════════════════════")
        for label, p in [("PRIMARY   (OpenAI)", self.openai),
                          ("SECONDARY (Gemini)", self.primary),
                          ("FALLBACK  (Groq)  ", self.fallback)]:
            if p:
                logging.warning("  %s  model=%-30s  key=✓", label, p["model"])
            else:
                logging.warning("  %s  NOT CONFIGURED", label)
        if not self.openai and not self.primary and not self.fallback:
            logging.warning("  ⚠ CRITICAL: no AI providers configured")
        logging.warning("════════════════════════════════════════════════════")


# Module-level singleton — used everywhere
_prov_mgr = AIProviderManager()

# Thin backwards-compat wrappers (used by BOQ import + sync AI helper)
def _make_chat_client(override_key=None):
    return _prov_mgr.best_client(override_key)

def _make_groq_client():
    p = _prov_mgr.fallback
    return (p["client"], p["model"]) if p else (None, None)

def _make_gemini_client(model=None):
    p = _prov_mgr.primary
    if not p:
        return None, None
    return p["client"], (model or p["model"])

# ── Odoo connection ────────────────────────────────────────────────────────────
# Set ODOO_API_KEY as a Railway environment variable (Variables tab) — never hardcode it.
ODOO_URL     = os.environ.get("ODOO_URL",     "https://alforat-beta.odoo.com").rstrip("/")
ODOO_DB      = os.environ.get("ODOO_DB",      "alforat-beta-35112787")
ODOO_UID     = int(os.environ.get("ODOO_UID", "2"))
ODOO_API_KEY = os.environ.get("ODOO_API_KEY", "")

logging.info("Odoo: %s  db=%s  uid=%d", ODOO_URL, ODOO_DB, ODOO_UID)

_odoo = xmlrpc.client.ServerProxy(ODOO_URL + "/xmlrpc/2/object")

# Multi-company context for construction company (company_id=3)
ODOO_CTX = {"allowed_company_ids": [1, 2, 3, 4, 5], "company_id": 3}

def odoo_call(model, method, args, kwargs=None):
    kw = dict(kwargs) if kwargs else {}
    if "context" not in kw:
        kw["context"] = ODOO_CTX
    return _odoo.execute_kw(ODOO_DB, ODOO_UID, ODOO_API_KEY,
                             model, method, args, kw)

# ── Tool definitions (OpenAI/Groq format) ─────────────────────────────────────
TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "odoo_search",
            "description": "Get records from Odoo (employees, units, contracts, projects, partners…).",
            "parameters": {
                "type": "object",
                "properties": {
                    "model":  {"type": "string", "description": "Odoo model. E.g. project.subcontracting.boq.line, boq.contract, construction.advance.payment, project.detailed.item.line, purchase.order, hr.employee"},
                    "domain": {"anyOf": [{"type": "string"}, {"type": "array"}], "description": "Filter domain. Use [] for all records. E.g. [[\"state\",\"=\",\"sale\"]]"},
                    "fields": {"type": "array", "items": {"type": "string"}, "description": "List of field names to return. E.g. [\"name\",\"state\",\"project_id\"]"},
                    "limit":  {"type": "integer", "description": "Max rows (default 50, max 200)"},
                    "order":  {"type": "string",  "description": "Sort. E.g. 'boq_cost desc'"}
                },
                "required": ["model", "domain", "fields"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "odoo_count",
            "description": "Count records in an Odoo model.",
            "parameters": {
                "type": "object",
                "properties": {
                    "model":  {"type": "string"},
                    "domain": {"anyOf": [{"type": "string"}, {"type": "array"}], "description": "Filter domain. Use [] for all records."}
                },
                "required": ["model", "domain"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "odoo_get_fields",
            "description": "List field names/types for an Odoo model. Use when unsure of field names.",
            "parameters": {
                "type": "object",
                "properties": {
                    "model": {"type": "string"}
                },
                "required": ["model"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "odoo_read_group",
            "description": "Group & aggregate Odoo records. Use for 'how many X per Y' or totals. Count is returned automatically as 'count' — never include 'id:count' in aggregates.",
            "parameters": {
                "type": "object",
                "properties": {
                    "model": {"type": "string", "description": "Odoo model name."},
                    "domain": {"anyOf": [{"type": "string"}, {"type": "array"}], "description": "Filter domain. Use [] for all records. E.g. [[\"state\",\"=\",\"draft\"]]"},
                    "groupby": {
                        "anyOf": [
                            {"type": "array", "items": {"type": "string"}},
                            {"type": "string"}
                        ],
                        "description": "Fields to group by. E.g. [\"project_id\"] or [\"state\"] or [\"partner_id\"]."
                    },
                    "aggregates": {
                        "anyOf": [
                            {"type": "array", "items": {"type": "string"}},
                            {"type": "string"}
                        ],
                        "description": "Numeric fields to sum. E.g. [\"boq_cost:sum\",\"quantity:sum\"]. Use [] if only need count. NEVER include 'id:count' — count is automatic."
                    }
                },
                "required": ["model", "domain", "groupby"]
            }
        }
    }
]

# ── Tool execution ─────────────────────────────────────────────────────────────
def _coerce_domain(raw):
    """Ensure domain is always a list, even when the LLM passes a string."""
    if isinstance(raw, list):
        return raw
    if isinstance(raw, str):
        raw = raw.strip()
        if raw in ("", "[]", "None", "null"):
            return []
        try:
            parsed = json.loads(raw)
            return parsed if isinstance(parsed, list) else []
        except Exception:
            return []
    return []

def _coerce_list(raw):
    """Ensure value is a list of strings, even when LLM passes a JSON string."""
    if isinstance(raw, list):
        return raw
    if isinstance(raw, str):
        raw = raw.strip()
        if raw in ("", "[]", "None", "null"):
            return []
        try:
            parsed = json.loads(raw)
            if isinstance(parsed, list):
                return parsed
            if isinstance(parsed, str):
                return [parsed]
        except Exception:
            pass
        # bare single field name like "project_id"
        return [raw] if raw else []
    return []

_MAX_RESULT_CHARS  = 12000  # max total chars returned to AI per tool call
_MAX_FIELD_CHARS   = 180    # max chars for any single string field value (long Arabic names)

# Correct Odoo's wrong Arabic accounting translations before sending to AI
_ACCOUNTING_TERM_MAP = {
    "مستصلح":       "مُسوَّاة",       # settled advance payment (Odoo mistranslates "settled" as "مستصلح")
    "مستصلحة":      "مُسوَّاة",
    "تسوية":        "مُسوَّاة",       # sometimes shown as تسوية
    "في التنفيذ":   "جارٍ",           # "run" state → جارٍ is more standard in Arabic accounting
    "مقفل":         "مُقفَّل",        # normalise spelling
}

def _fix_accounting_terms(val):
    """Replace Odoo's incorrect Arabic field translations with proper accounting terms."""
    if isinstance(val, str):
        return _ACCOUNTING_TERM_MAP.get(val, val)
    return val

def _trim_long_fields(rec):
    """Trim long strings and fix incorrect Odoo Arabic accounting translations."""
    if not isinstance(rec, dict):
        return rec
    out = {}
    for k, v in rec.items():
        if isinstance(v, str):
            v = _fix_accounting_terms(v)
            if len(v) > _MAX_FIELD_CHARS:
                v = v[:_MAX_FIELD_CHARS] + '…'
        out[k] = v
    return out

def _truncate_result(text):
    if len(text) > _MAX_RESULT_CHARS:
        return text[:_MAX_RESULT_CHARS] + f'\n... [truncated — {len(text)} chars total, showing first {_MAX_RESULT_CHARS}]'
    return text

def _flatten_m2o(val):
    """Convert Odoo many2one [id, "Name"] tuples to just the display name string."""
    if isinstance(val, (list, tuple)) and len(val) == 2 and isinstance(val[0], int) and isinstance(val[1], str):
        return _fix_accounting_terms(val[1])
    return val

def _flatten_record(rec):
    """Flatten all many2one fields in a record dict so the AI sees plain strings."""
    if isinstance(rec, dict):
        return {k: _flatten_m2o(v) for k, v in rec.items()}
    return rec

def run_tool(name, args):
    try:
        if name == "odoo_search":
            domain = _coerce_domain(args.get("domain", []))
            limit  = min(int(args.get("limit", 30)), 100)   # reduced: 30 default, 100 max
            raw_fields = args.get("fields", ["display_name"])
            fields = raw_fields if isinstance(raw_fields, list) else _coerce_list(raw_fields)
            if not fields:
                fields = ["display_name"]
            kwargs = {"fields": fields, "limit": limit}
            if args.get("order"):
                kwargs["order"] = args["order"]
            results = odoo_call(args["model"], "search_read", [domain], kwargs)
            flat = [_trim_long_fields(_flatten_record(r)) for r in results]
            return _truncate_result(json.dumps(flat, ensure_ascii=False, default=str))

        elif name == "odoo_count":
            domain = _coerce_domain(args.get("domain", []))
            count  = odoo_call(args["model"], "search_count", [domain])
            return json.dumps({"count": count})

        elif name == "odoo_get_fields":
            fields = odoo_call(args["model"], "fields_get", [],
                               {"attributes": ["string", "type"]})
            simple = {
                k: {"type": v["type"], "label": v["string"]}
                for k, v in fields.items()
                if v["type"] not in ("binary", "html", "serialized")
                and not k.startswith("message_")
                and not k.startswith("activity_")
            }
            return json.dumps(simple, ensure_ascii=False)

        elif name == "odoo_read_group":
            domain     = _coerce_domain(args.get("domain", []))
            groupby    = _coerce_list(args.get("groupby", []))
            agg_fields = _coerce_list(args.get("aggregates", []))
            # Strip invalid aggregates — Odoo count is automatic (__count), never pass id:count or bare 'count'
            agg_fields = [f for f in agg_fields if f not in ("id:count", "id", "__count", "count")]
            fields_list = list(groupby) + list(agg_fields)
            result = odoo_call(
                args["model"], "read_group",
                [domain, fields_list, groupby],
                {"lazy": False}
            )
            cleaned = []
            for row in result:
                item = {}
                for k, v in row.items():
                    if k == "__count":
                        item["count"] = v
                    elif k == "__domain":
                        continue
                    else:
                        v2 = _flatten_m2o(v)
                        item[k] = _fix_accounting_terms(v2) if isinstance(v2, str) else v2
                cleaned.append(item)
            return _truncate_result(json.dumps(cleaned, ensure_ascii=False, default=str))

    except Exception as e:
        err = str(e)
        if "Access Denied" in err or "Fault 3" in err:
            return json.dumps({
                "error": "Odoo Access Denied — the API key is expired or invalid.",
                "fix": "Go to Railway → Variables and update ODOO_API_KEY with a valid Odoo API key (Settings → Technical → API Keys in Odoo)."
            })
        return json.dumps({"error": err, "hint": "Try odoo_get_fields to check available fields"})

# ── Implementer tools: read tools + odoo_create ───────────────────────────────
IMPLEMENTER_TOOLS = TOOLS + [
    {
        "type": "function",
        "function": {
            "name": "odoo_create",
            "description": (
                "Create a new record in Odoo (project task, calendar event, internal note, etc.). "
                "ALWAYS call odoo_search first to resolve IDs (project_id, user_id, stage_id) "
                "before calling this tool — never guess integer IDs."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "model": {
                        "type": "string",
                        "description": "Odoo model name. E.g. project.task, calendar.event, note.note"
                    },
                    "values": {
                        "type": "object",
                        "description": (
                            "Field values as a JSON object. "
                            "project.task: {name, project_id (int), date_deadline (YYYY-MM-DD), description}. "
                            "calendar.event: {name, start (YYYY-MM-DD HH:MM:SS), stop, description}. "
                            "note.note: {name, memo}."
                        )
                    }
                },
                "required": ["model", "values"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "odoo_post_chatter",
            "description": (
                "Post a MOM (Minutes of Meeting) or internal note to the chatter of an Odoo record. "
                "Use this AFTER generating the MOM to save it permanently on a project task. "
                "ALWAYS call odoo_search first to find the task_id — never guess IDs."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "model":     {"type": "string",  "description": "Odoo model — usually 'project.task' or 'project.project'."},
                    "record_id": {"type": "integer", "description": "Integer ID of the record to post on."},
                    "body":      {"type": "string",  "description": "The MOM content (plain text or HTML)."}
                },
                "required": ["model", "record_id", "body"]
            }
        }
    }
]

def _run_implementer_tool(name, args):
    """Extended tool runner: read tools + odoo_create + odoo_post_chatter."""
    if name == "odoo_create":
        try:
            model  = args.get("model", "").strip()
            values = args.get("values", {})
            if not model:
                return json.dumps({"error": "model is required"})
            if not isinstance(values, dict) or not values:
                return json.dumps({"error": "values must be a non-empty object"})
            new_id = odoo_call(model, "create", [values])
            return json.dumps({"created_id": new_id, "model": model, "status": "ok"})
        except Exception as e:
            err = str(e)
            if "Access Denied" in err or "Fault 3" in err:
                return json.dumps({"error": "Odoo Access Denied — API key may lack write permissions"})
            return json.dumps({"error": err})

    if name == "odoo_post_chatter":
        try:
            model     = args.get("model", "project.task").strip()
            record_id = int(args.get("record_id", 0))
            body      = args.get("body", "").strip()
            if not record_id or not body:
                return json.dumps({"error": "record_id and body are required"})
            odoo_call(model, "message_post", [[record_id]], {
                "body": body,
                "message_type": "comment",
                "subtype_xmlid": "mail.mt_note"
            })
            return json.dumps({"status": "ok", "model": model, "record_id": record_id,
                               "message": "MOM saved to Odoo chatter successfully"})
        except Exception as e:
            err = str(e)
            if "Access Denied" in err or "Fault 3" in err:
                return json.dumps({"error": "Odoo Access Denied — API key may lack write permissions"})
            return json.dumps({"error": err})

    return run_tool(name, args)

# ── System prompt ──────────────────────────────────────────────────────────────
def get_system_prompt():
    from datetime import date as _date
    today_date = _date.today().isoformat()   # e.g. "2026-08-05"
    return f"""You are an AI business assistant connected to an Odoo ERP system (Real Estate, Construction, HR, Procurement, Inventory). Reply in the same language as the user (Arabic or English).

TODAY = {today_date}  ← use this exact string for any date comparison domains (e.g. [['date_to','<','{today_date}']])

══ SCOPE — STEP-BY-STEP CHECK (do this BEFORE deciding to refuse) ══

STEP 1 — ALWAYS check for business keywords first. If the question contains ANY of these, it is IN-SCOPE → call a tool immediately, NEVER refuse:
  Arabic business terms: مقاول / مقاولين / باطن / مدفوعات / دفعات / سلف / عقد / عقود / مشروع / مشاريع /
    موظف / موظفين / مخزون / مستودع / مشتريات / وحدة / شقة / عقار / BOQ / مقايسة / فاتورة / دفعة /
    مقارنة / إجمالي / قيمة / كمية / تكلفة / ميزانية / تقرير / حالة / مرحلة / خطة / تأخير / متأخر /
    راتب / رواتب / مرتب / حضور / غياب / إجازة / توظيف / تقييم / أداء / مؤشر / عقد عمل / كشف رواتب /
    مرشح / دوام / انصراف / قسيمة / تجربة / مغادرة / مهارة / درجة / بدل / خصم /
    رسم بياني / مخطط / عرض / قسم / إدارة / اعرض / أظهر / قائمة / عدد / تجميع
  English business terms: project / employee / contractor / advance / payment / contract / BOQ / inventory /
    warehouse / purchase / unit / invoice / budget / cost / quantity / report / status / phase / plan / overdue /
    payroll / salary / attendance / leave / hire / appraisal / performance / KPI / recruitment / probation /
    headcount / payslip / timeoff / onboarding / offboarding / department / vacancy / skill / overtime /
    chart / graph / bar chart / pie chart / show / display / list / count / grouped / breakdown / summary

STEP 2 — Only if the question contains NONE of the above terms, check if it's truly off-topic:
  OUT-OF-SCOPE (refuse only these): weather, cooking, sports, jokes, history, geography, personal advice, general coding unrelated to Odoo

STEP 3 — When refusing, respond ONLY with:
• Arabic: "أنا مساعد أعمال متخصص في بيانات Odoo فقط. يمكنني مساعدتك في المشاريع، العقارات، الموظفين، المشتريات، أو المخزون."
• English: "I'm a business assistant specialized in Odoo data only. I can help you with projects, real estate, employees, procurement, or inventory."

⚠ CRITICAL: Questions like "كم إجمالي المدفوعات المقدمة لكل مقاول مقارنةً بقيمة عقده؟" are ALWAYS IN-SCOPE — call tools immediately.
⚠ CRITICAL: ANY request that includes "chart", "graph", "bar", "pie", "show", "list", "display", "count" about Odoo entities IS IN-SCOPE — the frontend renders charts automatically. NEVER refuse these. Just fetch the data using tools and return it as a markdown table — the UI handles the visualization.
⚠ CRITICAL: "Show employee count by department with a bar chart" = IN-SCOPE → call odoo_read_group on hr.employee grouped by department_id. NEVER refuse.

══ ARABIC BUSINESS TERMINOLOGY — never treat these as search values ══
These are CATEGORY NAMES / CONCEPTS, not actual record names. Map them to the correct Odoo model:
• مقاولي الباطن / مقاول الباطن / الباطن      → subcontractor.contract (partner_id = actual contractor name)
• بى اى كيو / BOQ / مقايسة / بنود الأعمال    → project.subcontracting.boq.line OR project.detailed.item.line
• المدفوعات المقدمة / الدفعات المقدمة / السلف → construction.advance.payment
• العقود / عقود المقاولين                     → subcontractor.contract
• الموظفين / الموظف                           → hr.employee
• المشاريع / المشروع                          → project.project
• الوحدات العقارية / الشقق / العقارات         → rs.dev.unit
• المشتريات / أوامر الشراء                   → purchase.order
• المخزون / المستودع / الكميات               → stock.quant (with location_id.usage=internal filter)

⚠ CRITICAL — NEVER do this:
- NEVER search partner_id or name = "مقاولي الباطن" — this is the MODEL TYPE, not a contractor name
- NEVER use Arabic category words as domain filter values
- If user says "مقاولي الباطن" they mean ALL subcontractors — use domain=[] on subcontractor.contract
- If user says "بى اى كيو" they mean BOQ records — query the BOQ model, do NOT search for a partner named "بى اى كيو"

══ QUERY VALIDATION — before executing any search ══
- Re-read the question and identify: WHAT data (model) + WHICH filter (domain) + WHAT fields
- If the question mentions a business term (مقاولي الباطن, BOQ, etc.) → map it to the correct model first
- If a name/value seems to be a category concept rather than a real record name → do NOT use it as a domain filter value
- Use domain=[] to get ALL records when user asks for "كل" (all) without specifying a particular name

RULES:
- CRITICAL: You MUST call a tool for EVERY question about data. NEVER answer data questions from memory or imagination — the data changes daily. If you answer without calling a tool, your answer WILL be wrong.
- Use tools for real data. Never guess or fabricate numbers, names, statuses, or values.
- BATCH: Call ALL needed tools in ONE response — never chain across turns.
- After tool results arrive, answer immediately. Do NOT call more tools.
- NEVER call the same model more than once per turn — pick the right fields the first time.
- Totals/counts → odoo_read_group. Lists → odoo_search (limit 50, increase to 200 for "all records").
- Format numbers with commas. Currency = EGP. Use markdown tables.
- On field error → odoo_get_fields once, then retry. Multi-company context is automatic.
- If a tool returns an error, report the error — do NOT invent data to compensate.

CHARTS: CHART_BAR:{{"title":"T","labels":["A","B"],"data":[10,20]}}  CHART_PIE:{{"title":"T","labels":["A"],"data":[1]}}

══ CONSTRUCTION MODELS — use ONLY these for any construction/project question ══

project.project  →  construction projects
  Fields: name, user_id, partner_id, date_start, date (planned end),
          expected_end_date, actual_end_date, state, last_update_status,
          total_planned_cost, total_actual_cost
  NOTE: fetch ALL these fields in ONE call — never call project.project twice

main.project.plan.operation  →  construction plan milestones (top-level grouping)
  Fields: name, project_id, plan_type, user_id, sub_plan_count
  plan_type values: subcontract | raw_material | assets | labor | misc
  ⚠ WARNING: the "date" field = system creation timestamp (auto-set = same as create_date) — NOT a user-set deadline
  ⚠ NEVER use "date" on this model to determine behind-schedule status — it is meaningless for scheduling
  NOTE: NO state/progress/completion field — cannot determine done/overdue from this model alone
  → For "behind schedule" milestones: query sub.project.plan.operation and use main_plan_id to identify parent

sub.project.plan.operation  →  sub-operations under a milestone — ONLY model with real date ranges
  Fields: name, project_id, main_plan_id, date_from (Period From), date_to (Period To), plan_type, user_id
  date_from / date_to = user-set planned work period dates (YYYY-MM-DD format, set by project managers)
  ▶ "Behind schedule / overdue" = date_to < TODAY ({today_date}) AND date_to != False
  Domain for overdue: [{{"date_to","<","{today_date}"}},{{"date_to","!=",False}}]
  NOTE: NO state/progress/completion field on this model either

project.detailed.item.line  →  client BOQ line items
  Fields: name, project_id, quantity, unit_cost, total_cost (=Total Price), main_item_line_id, boq_item_id

project.subcontracting.boq.line  →  raw subcontractor BOQ items (scope definition only)
  Fields: name, project_id, quantity, unit_cost, billed_qty, remain_qty
  WARNING: boq_cost = always 0, do NOT use for financial totals
  ⚠ This model has NO direct link to the subcontractor (partner) name
  → To show subcontractor name alongside BOQ items, use subcontractor.contract.line instead (see below)

project.main.item.line  →  BOQ section/category headers
  Fields: name, project_id

subcontractor.contract  →  subcontractor contracts — PRIMARY financial model
  Fields: name, project_id, partner_id (= subcontractor company name), contract_value (=THE contract total),
          bills_amount_total, bills_amount_due, total_adv_amount
  status field values: draft | running | closed   (field name is "status" NOT "state")

subcontractor.contract.line  →  assigned BOQ line items inside a contract — USE THIS for subcontractor BOQ with partner name
  Fields: name, contract_id, project_id, product_id (= item description), unit_price, total_price, billed_qty, remain_qty, boq_id
  RELATIONSHIP: contract_id → subcontractor.contract.partner_id = subcontractor name
  ▶ For "subcontractor BOQ items + who is the subcontractor": call BOTH in one turn:
    1) odoo_search subcontractor.contract.line  fields=[name,project_id,product_id,quantity,unit_price,total_price,contract_id,billed_qty,remain_qty]
    2) odoo_search subcontractor.contract       fields=[name,project_id,partner_id,contract_value,status]
    Then join on contract_id = contract.name to show the subcontractor (partner_id) for each line

boq.contract  →  BOQ scope config record — EMPTY in database (0 records), NO financial data
  Fields: name, project_id, partner_id
  ⚠ NEVER use boq.contract for any financial or contract-value queries — it has no money fields and no data

construction.advance.payment  →  advance payments to contractors/clients
  Fields: name, partner_id, amount, state, project_id, due_amount, settled_amount, payment_type
  state values: draft | run | settled   (NOT "running")
  payment_type values: customer | subcontractor

project.task  →  project tasks
  Fields: name, project_id, stage_id, date_deadline, kanban_state

project.boq.item  →  global BOQ item catalogue
  Fields: name, code, uom_id, billing_type (billable|unbillable)

══ REAL ESTATE MODELS — use ONLY for RE/property questions, NEVER for construction ══

rs.dev.unit  →  RE property units for sale (USE THIS, not project.rs.unit)
  Fields: unit_code, rs_project_id, rs_building_id, phase_id, unit_price,
          sold_price, net_area, paid_amount, price_after_discount,
          usability_type (commercial|administrative|residential|medical|parking_slot),
          unit_type (apartment|duplex|villa|twin_house|town_house),
          historical_process (holded|booked|reserved|contracted)

project.rs.building  →  RE buildings (name, project_id)
project.rs.phase     →  RE development phases (name, project_id) ← NOT for construction

══ INVENTORY MODELS — use ONLY for stock/warehouse/inventory questions ══

stock.quant  →  current stock on hand
  Fields: product_id, location_id, quantity, reserved_quantity
  ⚠ CRITICAL: stock.quant stores BOTH real stock AND virtual (negative) counterparts.
  ALWAYS add domain [["location_id.usage","=","internal"]] to show only real stock.
  Without this filter, sum(quantity) = 0 because virtual locations cancel real ones.
  ▶ For "current stock / stock levels": domain=[["location_id.usage","=","internal"],["quantity",">",0]]
  ▶ For "stock by product": odoo_read_group stock.quant domain=[["location_id.usage","=","internal"]] groupby=["product_id"] aggregates=["quantity:sum"]
  ▶ For "stock by location": odoo_read_group stock.quant domain=[["location_id.usage","=","internal"]] groupby=["location_id"] aggregates=["quantity:sum"]

stock.picking  →  stock transfers (receipts, deliveries, internal)
  Fields: name, partner_id, state, picking_type_id, scheduled_date, date_done, origin
  state values: draft | waiting | confirmed | ready (= assigned) | done | cancel
  picking_type_id.code values: incoming (receipt) | outgoing (delivery) | internal
  ▶ Pending receipts: domain=[["picking_type_id.code","=","incoming"],["state","in",["confirmed","assigned"]]]
  ▶ Overdue deliveries: domain=[["picking_type_id.code","=","outgoing"],["state","not in",["done","cancel"]],["scheduled_date","<","TODAY"]]

stock.warehouse  →  warehouses
  Fields: name, code, lot_stock_id, wh_input_stock_loc_id, wh_output_stock_loc_id

stock.location  →  storage locations
  Fields: name, complete_name, usage, location_id
  usage values: internal | view | supplier | customer | inventory | production | transit

stock.lot  →  lot/serial numbers for traceability
  Fields: name, product_id, expiration_date, product_qty, ref
  ▶ Expiring lots: domain=[["expiration_date","!=",False],["expiration_date","<","DATE+30DAYS"]]

══ INVENTORY ROUTING ══
• stock levels / مخزون / كميات                           → stock.quant ALWAYS with [["location_id.usage","=","internal"]]
• transfers / receipts / deliveries / نقل / استلام        → stock.picking
• warehouses / مستودعات                                  → stock.warehouse
• locations / مواقع تخزين                                → stock.location where usage='internal'
• lot / serial / دفعات / أرقام تسلسلية                  → stock.lot

══ SHARED MODELS ══
purchase.order      → purchase orders
  Fields: name, partner_id, state, amount_total, date_order
  state values: draft=RFQ | sent=RFQ Sent | to approve=To Approve | purchase=Purchase Order | done=Locked | cancel=Cancelled

purchase.order.line → PO lines (order_id,product_id,product_qty,price_unit,price_subtotal)
res.partner         → partners/vendors/clients (name,phone,email,is_company)

══ HR MODELS — use these for ALL HR / people management questions ══

FULL HR LIFECYCLE: Requisition → Recruitment → Onboarding → Contract → Attendance → Leaves → Payroll → Appraisal/PMS → Probation Review → Employee Movement → Offboarding

── EMPLOYEES ──
hr.employee  →  employee master data
  Fields: name, employee_number, department_id, job_id, job_title, parent_id (manager),
          coach_id, work_phone, work_email, work_location_id, active,
          gender, marital, birthday, country_id, resource_calendar_id, company_id
  ▶ Active employees: domain=[["active","=",True]]
  ▶ Inactive / offboarded: domain=[["active","=",False]]
  ▶ Headcount by dept: odoo_read_group hr.employee groupby=["department_id"] aggregates=[]
  NOTE: When user asks for employees, ALWAYS include active=True in domain unless they ask for inactive

── DEPARTMENTS & ORG STRUCTURE ──
hr.department  →  departments / org chart
  Fields: name, manager_id, parent_id, member_count, active, company_id
  ▶ Root departments: domain=[["parent_id","=",False]]
  ▶ Sub-departments: domain=[["parent_id","=",<dept_id>]]

hr.job  →  job positions / headcount plan / vacancies
  Fields: name, department_id, no_of_recruitment (open slots), no_of_hired_employee,
          no_of_employee (current headcount), state, description
  state: recruit (open/hiring) | open (filled)
  ▶ Positions currently recruiting: domain=[["state","=","recruit"]]

── RECRUITMENT ──
hr.applicant  →  job applications / hiring pipeline
  Fields: partner_name, email_from, partner_phone, job_id, department_id,
          stage_id, user_id (recruiter), priority, date_open, date_closed,
          kanban_state, source_id, salary_expected, salary_proposed
  priority: 0=Normal | 1=Good | 2=Very Good | 3=Excellent
  ▶ Active pipeline: domain=[["active","=",True]]
  ▶ By stage: odoo_read_group groupby=["stage_id"] aggregates=[]
  ▶ By job: odoo_read_group groupby=["job_id"] aggregates=[]
  ▶ By source: odoo_read_group groupby=["source_id"] aggregates=[]

── CONTRACTS ──
hr.contract  →  employment contracts
  Fields: name, employee_id, job_id, department_id, wage, contract_type_id,
          date_start, date_end, state, structure_type_id, hr_responsible_id, notes
  state: draft | open | close | cancel
  ▶ Active contracts: domain=[["state","=","open"]]
  ▶ Expiring in 30 days: domain=[["state","=","open"],["date_end","!=",False],["date_end","<","{today_date}"]]
  ▶ Probation contracts: domain=[["state","=","open"],["contract_type_id.name","ilike","probation"]]
  ▶ By dept: odoo_read_group groupby=["department_id"] aggregates=["wage:sum"]

── ATTENDANCE ──
hr.attendance  →  daily check-in / check-out records
  Fields: employee_id, check_in, check_out, worked_hours
  ▶ Today's records: domain=[["check_in",">=","{today_date} 00:00:00"]]
  ▶ Missing check-out: domain=[["check_out","=",False]]
  ▶ Monthly hours per employee: odoo_read_group groupby=["employee_id"] aggregates=["worked_hours:sum"]
    + add date domain for the month range
  NOTE: check_in / check_out stored in UTC — displayed in user timezone in Odoo UI

── LEAVES / TIME OFF ──
hr.leave  →  leave requests (individual time-off records)
  Fields: employee_id, holiday_status_id (leave type), date_from, date_to,
          number_of_days, state, department_id, name
  state: draft | confirm | validate1 | validate | refuse
  ▶ Approved leaves: domain=[["state","=","validate"]]
  ▶ Pending approval: domain=[["state","in",["confirm","validate1"]]]
  ▶ By leave type: odoo_read_group groupby=["holiday_status_id"] aggregates=["number_of_days:sum"]
  ▶ By department: odoo_read_group groupby=["department_id"] aggregates=["number_of_days:sum"]

hr.leave.allocation  →  leave balance allocations per employee
  Fields: employee_id, holiday_status_id, number_of_days, state, date_from, date_to
  state: draft | confirm | validate1 | validate | refuse
  ▶ Approved allocations: domain=[["state","=","validate"]]
  ▶ By employee: odoo_read_group groupby=["employee_id"] aggregates=["number_of_days:sum"]

hr.leave.type  →  leave type / policy definitions
  Fields: name, allocation_type, max_leaves, leaves_taken, remaining_leaves

── PAYROLL ──
hr.payslip  →  individual payslips
  Fields: employee_id, date_from, date_to, state, struct_id, payslip_run_id, name
  state: draft | verify | done | cancel
  ▶ Paid payslips: domain=[["state","=","done"]]
  ▶ Current month (paid): domain=[["date_from",">=","{today_date[:7]}-01"],["date_from","<","{today_date}"],["state","=","done"]]
  ▶ By employee: odoo_read_group groupby=["employee_id"] aggregates=[]
  ▶ By run: odoo_read_group groupby=["payslip_run_id"] aggregates=[]

hr.payslip.run  →  payroll batches / pay runs
  Fields: name, date_start, date_end, state
  state: draft | verify | close
  ▶ Closed / disbursed: domain=[["state","=","close"]]

── APPRAISAL / PERFORMANCE MANAGEMENT ──
hr.appraisal  →  ALL appraisal records (standard + EPA custom module — same table)
  Fields: employee_id, department_id, company_id, batch_id, evaluation_period_id,
          state, date_close, date_deadline,
          final_grade, epa_final_grade, result_grade,
          total_score, final_score, epa_final_score, kpi_total_score, kpi_total_weight,
          is_employee_done, is_manager_done, is_senior_done, hr_delivered,
          manager_feedback, employee_feedback, development_plan, appraisal_purpose_id

  ⚠ COMPLETE WORKFLOW — 8 real state values (NOT "done", never use "done"):
  new               → Just created, not yet started
  pending           → Submitted / waiting to begin
  self_evaluation   → Employee filling their self-assessment
  manager_evaluation → Direct manager evaluating
  senior_evaluation  → Senior manager evaluating
  hr_evaluation      → HR department final review
  completed          → Fully completed (use "completed", NOT "done")
  cancel             → Cancelled

  Grade fields (for completed appraisals):
  - final_grade:     ++A | +A | A | -A | B | … (overall performance grade)
  - epa_final_grade: A++ | A+ | A | A- | B+ | B | D (EPA grading scale)
  - result_grade:    pp_a=++A | p_a=+A | a=A | m_a=-A | c=C (internal enum)

  ▶ All appraisals by status:  odoo_read_group groupby=["state"] aggregates=[]
  ▶ Completed:                 domain=[["state","=","completed"]]
  ▶ In-progress (any step):    domain=[["state","in",["self_evaluation","manager_evaluation","senior_evaluation","hr_evaluation"]]]
  ▶ Not yet started:           domain=[["state","in",["new","pending"]]]
  ▶ Cancelled:                 domain=[["state","=","cancel"]]
  ▶ By department:             odoo_read_group groupby=["department_id"] aggregates=[]
  ▶ Avg scores (completed):    domain=[["state","=","completed"]] then odoo_read_group groupby=["department_id"] aggregates=["total_score:avg","epa_final_score:avg","kpi_total_score:avg"]
  ▶ By grade (completed):      odoo_read_group groupby=["final_grade"] aggregates=[] domain=[["state","=","completed"]]
  ▶ By period:                 domain=[["evaluation_period_id","=",<period_id>]]
  ▶ By batch:                  domain=[["batch_id","=",<batch_id>]]

  ⚠ EMPTY RESULT RULE: If a state-filtered query returns 0 records, you MUST immediately
    also call odoo_read_group on hr.appraisal groupby=["state"] aggregates=[] (no domain)
    to show how many appraisals exist in each stage. Present the by-state breakdown in
    your answer so the user understands what stage their appraisals are currently at.
    Example: "No completed appraisals found. Current status: 3 in manager_evaluation, 2 in self_evaluation, 1 in hr_evaluation."

hr.appraisal.review  →  Multi-evaluator review entries linked to an appraisal
  Each appraisal can have multiple reviewers; this stores each reviewer's submission.
  Fields: appraisal_id, evaluator_id, evaluator_type, status, score, overall_comment, submitted_date
  evaluator_type: employee | direct_manager | senior_manager | hr
  status: pending | in_progress | submitted
  ⚠ Uses "status" NOT "state"
  ▶ Submitted reviews:  domain=[["status","=","submitted"]]
  ▶ Pending reviews:    domain=[["status","in",["pending","in_progress"]]]
  ▶ By evaluator type:  odoo_read_group groupby=["evaluator_type"] aggregates=[]
  ▶ By evaluator type + status: odoo_read_group groupby=["evaluator_type","status"] aggregates=[]

hr.appraisal.review.line  →  Individual criterion scores within a review
  Fields: review_id, criterion_id, section_name, criterion_name, rating, comment, weight_pct, weighted_score
  ▶ By section: odoo_read_group groupby=["section_name"] aggregates=["weighted_score:sum"]

hr.appraisal.line  →  Appraisal section/criteria scoring lines (per appraisal)
  Fields: appraisal_id, section_id, criterion_id, section_name, criterion_name,
          score, score_float, section_weight, weighted_score, is_mandatory, assessor_done
  ▶ Scores for one appraisal: domain=[["appraisal_id","=",<id>]]
  ▶ Avg by section: odoo_read_group groupby=["section_name"] aggregates=["weighted_score:avg"]

hr.appraisal.goal  →  Employee development goals linked to appraisals
  Fields: appraisal_id, dev_goal_appraisal_id, employee_id, manager_id,
          name, progression, deadline, goal_status, priority, goal_type,
          weight, kpi_score, weighted_score, kpi_grade
  goal_status: not_started | in_progress | completed (use goal_status NOT state)
  priority: 0=Normal | 1=High | 2=Critical
  ▶ All goals: domain=[]
  ▶ By status: odoo_read_group groupby=["goal_status"] aggregates=[]
  ▶ Overdue: fetch all with deadline < today then filter goal_status != completed
  ▶ By employee: domain=[["employee_id","=",<emp_id>]]

epa.appraisal.batch  →  Appraisal batch / cycle management records
  ⚠ "epa.appraisal" does NOT exist — use "epa.appraisal.batch" for batches.
  Fields: name, period_id, department_id, manager_id, appraisal_template_id, state, date_close, appraisal_count
  state: draft | in_progress
  ▶ Active batches: domain=[["state","=","in_progress"]]
  ▶ All batches: domain=[]

epa.period  →  Evaluation periods / assessment cycles
  Fields: name, state, date_from, date_to, notes, active
  state: draft | open | closed
  ▶ Open periods: domain=[["state","=","open"]]
  ▶ All periods: domain=[]
  ▶ Current period: domain=[["state","=","open"],["date_from","<=","{today_date}"],["date_to",">=","{today_date}"]]

hr.performance.evaluation  →  Standalone performance evaluation forms (separate from hr.appraisal)
  Fields: name, employee_id, evaluator_id, senior_mgr_id, period_id, template_id,
          state, final_grade, final_score, kpi_section_score, behaviour_section_score, mgmt_section_score,
          employee_comments, manager_notes, submission_date
  state: draft | pending_employee | employee_submitted | manager_submitted | senior_approved | closed
  ▶ Closed evaluations: domain=[["state","=","closed"]]
  ▶ By state: odoo_read_group groupby=["state"] aggregates=[]
  ▶ Score summary: odoo_read_group groupby=["period_id"] aggregates=["final_score:avg"]

hr.performance.kpi  →  KPI LIBRARY / DEFINITIONS (NOT per-employee results)
  This is the catalogue of KPI definitions — NOT individual employee scores.
  Fields: name, code, unit, calculation_method, default_weight, target, state, requested_by, department_id, active
  state: draft | pending | approved | rejected
  ⚠ NO employee_id, NO achievement, NO period_id, NO kpi_group_id on this model.
  ▶ All approved KPI definitions: domain=[["state","=","approved"]]
  ▶ Use this ONLY to list/browse KPI definitions — never for employee scores.

hr.appraisal.kpi.result  →  Per-employee KPI RESULTS linked to an appraisal (custom module)
  This is the correct model for "who achieved their KPI / نسبة تحقيق مؤشرات الأداء".
  Fields: employee_id, appraisal_id, kpi_id, kpi_display_name, weight, target, actual,
          achievement_pct, weighted_score, kpi_grade, notes, unit
  achievement_pct = (actual / target × 100) — already computed, stored as a number 0–100+
  ▶ All results: domain=[]
  ▶ Achieved target (≥100%): fetch all then filter where achievement_pct >= 100
  ▶ Avg achievement by employee: odoo_read_group groupby=["employee_id"] aggregates=["achievement_pct:avg","weighted_score:avg"]
  ▶ Results for one appraisal: domain=[["appraisal_id","=",<id>]]
  ▶ Results for one employee: domain=[["employee_id","=",<emp_id>]]
  ▶ Summary across all: odoo_read_group groupby=["employee_id"] aggregates=["achievement_pct:avg","weight:sum"]
  NOTE: achievement_pct is a PERCENTAGE — show as average (متوسط), never as sum.

hr.appraisal.kpi  →  Arabic-form KPI lines (performance evaluation form)
  Fields: appraisal_id, name, name_ar, weight, target_value, actual_value, score, weighted_score
  ▶ By appraisal: domain=[["appraisal_id","=",<id>]]
  ▶ Avg score by appraisal: odoo_read_group groupby=["appraisal_id"] aggregates=["score:avg","weighted_score:sum"]

── SKILLS ──
hr.employee.skill  →  employee skills / competency matrix
  Fields: employee_id, skill_id, skill_level_id, skill_type_id, level_progress
  ▶ By skill type: odoo_read_group groupby=["skill_type_id"] aggregates=[]
  ▶ By employee: domain=[["employee_id","=",<emp_id>]]

══ CUSTOM HR REPORTS SUMMARY ══
| Report                  | Model                | Key Filter                   | Group By          |
|-------------------------|----------------------|------------------------------|-------------------|
| Headcount by Dept       | hr.employee          | active=True                  | department_id     |
| Employee Directory      | hr.employee          | active=True                  | job_id            |
| Open Vacancies          | hr.job               | state=recruit                | department_id     |
| Recruitment Pipeline    | hr.applicant         | active=True                  | stage_id          |
| Hire Source Analysis    | hr.applicant         | date_closed (month)          | source_id         |
| Active Contracts        | hr.contract          | state=open                   | department_id     |
| Expiring Contracts      | hr.contract          | state=open, date_end<+30d    | employee_id       |
| Probation Due           | hr.contract          | type=probation, state=open   | date_end          |
| Daily Attendance        | hr.attendance        | check_in >= today            | employee_id       |
| Monthly Hours           | hr.attendance        | date range (worked_hours sum)| employee_id       |
| Pending Leave Requests  | hr.leave             | state=confirm                | department_id     |
| Leaves by Type          | hr.leave             | state=validate, month        | holiday_status_id |
| Leave Allocations       | hr.leave.allocation  | state=validate               | employee_id       |
| Payroll Batches         | hr.payslip.run       | state=close                  | date_start        |
| Monthly Payslips        | hr.payslip           | state=done, month            | employee_id       |
| Appraisal by Status     | hr.appraisal         | —                            | state             |
| Appraisal by Dept       | hr.appraisal         | —                            | department_id     |
| Appraisal Scores        | hr.appraisal         | state=completed              | department_id     |
| Appraisal by Grade      | hr.appraisal         | state=completed              | final_grade       |
| Pending Reviews         | hr.appraisal.review  | status=pending/in_progress   | evaluator_type    |
| Development Goals       | hr.appraisal.goal    | —                            | goal_status       |
| Evaluation Periods      | epa.period           | state=open                   | date_from         |
| Appraisal Batches       | epa.appraisal.batch  | state=in_progress            | department_id     |
| Perf. Evaluations       | hr.performance.evaluation | state=closed            | period_id         |
| KPI Achievement         | hr.appraisal.kpi.result | —                         | employee_id       |
| KPI Definitions         | hr.performance.kpi   | state=approved               | department_id     |
| Employee Skills Matrix  | hr.employee.skill    | —                            | skill_type_id     |

══ ROUTING — which model to answer each question ══
• projects / list of projects / مشاريع                       → project.project (NEVER project.rs.phase)
• plan / milestones / schedule / خطة / مراحل                 → main.project.plan.operation + sub.project.plan.operation
• behind schedule / overdue / متأخر / delay / تأخير          → sub.project.plan.operation domain [{{"date_to","<","{today_date}"}},{{"date_to","!=",False}}]
                                                               then identify parent milestone via main_plan_id field
• client BOQ / مقايسة العميل                                 → project.detailed.item.line (total_cost for amounts)
• subcontractor BOQ items / بنود الباطن (scope only, no partner)  → project.subcontracting.boq.line
• subcontractor BOQ items WITH subcontractor name / مع المقاول → subcontractor.contract.line + subcontractor.contract (join on contract_id)
• BOQ contracts / subcontractor contracts / عقود المقاولين   → subcontractor.contract (use contract_value)
• contract value / contract total / قيمة العقد               → subcontractor.contract (use contract_value)
• advance payments / مدفوعات مقدمة                          → construction.advance.payment
• RE units / وحدات عقارية / شقق / عقارات                    → rs.dev.unit
• purchases / مشتريات                                       → purchase.order
• employees / headcount / موظفين / عدد الموظفين            → hr.employee  (domain=[["active","=",True]] unless inactive asked)
• departments / org structure / إدارات / هيكل تنظيمي       → hr.department
• job positions / vacancies / وظائف / شواغر                → hr.job
• recruitment / applicants / pipeline / توظيف / مرشحين     → hr.applicant
• employment contracts / عقود العمل                         → hr.contract
• attendance / check-in / حضور / انصراف / دوام             → hr.attendance
• leaves / time-off / إجازات / غياب                        → hr.leave  (allocations → hr.leave.allocation)
• payroll / salary / payslips / رواتب / مرتبات / كشف راتب  → hr.payslip + hr.payslip.run
• appraisal status / تقييمات حسب الحالة / مكتمل / معلق     → hr.appraisal groupby=["state"] (8 states: new/pending/self_evaluation/manager_evaluation/senior_evaluation/hr_evaluation/completed/cancel)
• appraisal scores / درجات التقييم / نتائج                 → hr.appraisal domain=[["state","=","completed"]] fields=[total_score,epa_final_score,final_grade]
• appraisal reviews / مراجعات / multi-evaluator             → hr.appraisal.review (status: pending/in_progress/submitted)
• development goals / أهداف التطوير / dev goals             → hr.appraisal.goal (goal_status not state)
• evaluation periods / دورات التقييم / فترات                → epa.period (state: draft/open/closed)
• appraisal batches / دفعات التقييم                         → epa.appraisal.batch (state: draft/in_progress)
• performance evaluation / تقييم الأداء (standalone form)   → hr.performance.evaluation (state: draft/pending_employee/employee_submitted/manager_submitted/senior_approved/closed)
• ⚠ epa.appraisal DOES NOT EXIST — never use it; use hr.appraisal for all appraisal queries
• KPI achievement / نسبة تحقيق / من حقق / employee scores  → hr.appraisal.kpi.result  (has achievement_pct, employee_id)
• KPI definitions / KPI library / مكتبة المؤشرات          → hr.performance.kpi  (definitions only, NO scores)
• skills / competency / مهارات الموظفين                    → hr.employee.skill
• offboarding / resigned / departed / مغادرة / مستقيل      → hr.employee domain=[["active","=",False]]

CRITICAL RULES:
- project.rs.phase and project.rs.unit are REAL ESTATE ONLY — NEVER for construction
- rs.dev.unit is the correct RE sales model (project.rs.unit has NO price/status data)
- For ALL contract/financial queries: use subcontractor.contract with contract_value field
- boq.contract is EMPTY (no records) — NEVER use it for any query, it will return nothing
- construction.advance.payment state = "run" (not "running")
- COUNTING: odoo_read_group aggregates=[] gives count automatically; never add id:count
- Domain cannot compare two fields — fetch rows and filter in the answer text
- Call each model ONLY ONCE per turn — include all needed fields in a single call
- Many2one fields return as plain strings (e.g. project_id = "INFINITY MALL - Phase 1") — use them directly in the table
- BEHIND SCHEDULE: ONLY sub.project.plan.operation has real date ranges; use domain [{{"date_to","<","{today_date}"}},{{"date_to","!=",False}}]
  main.project.plan.operation.date = creation timestamp only — NEVER use it for overdue/schedule checks"""

def get_implementer_system_prompt():
    from datetime import date as _date
    today_date = _date.today().isoformat()
    return f"""You are an Odoo Implementation Partner — an AI that runs meetings, captures minutes of meeting (MOM), and creates action items directly in the Odoo ERP system. Reply in the same language as the user (Arabic or English).

TODAY = {today_date}

══ YOUR ROLE ══
You are NOT a read-only query assistant. You can:
1. SEARCH Odoo data to look up projects, employees, tasks
2. CREATE records: project tasks (project.task), calendar events (calendar.event)
3. Generate structured bilingual MOM (Minutes of Meeting)
4. Track meeting decisions and action items

══ MEETING FLOW ══
When user says "start meeting" / "ابدأ اجتماع" / "بدأ الاجتماع":
- Acknowledge: tell user recording has started
- Ask (once): meeting topic + attendees names + related project (if any)
- Respond concisely — do NOT ask multiple clarification questions

During meeting (user sends notes, decisions, action items):
- Acknowledge each input briefly
- Keep track of what was said for the MOM

When user says "generate MOM" / "استخرج المحضر" / "end meeting" / "إنهاء الاجتماع":
- Output a full structured MOM in this exact format (bilingual):

---
## 📋 محضر اجتماع | Minutes of Meeting
**التاريخ / Date:** {today_date}
**الموضوع / Topic:** [meeting topic]
**الحضور / Attendees:** [list of names]

### القرارات | Decisions
1. [decision 1]
2. [decision 2]

### بنود العمل | Action Items
| # | المهمة / Task | المسؤول / Owner | الموعد / Deadline |
|---|---|---|---|
| 1 | [task description] | [owner name] | [YYYY-MM-DD] |

### ملاحظات | Notes
[any additional notes]

---

After generating MOM, ask: "هل تريد حفظ المحضر في Odoo وإنشاء المهام؟ / Should I save the MOM to Odoo and create the tasks?"

══ SAVING MOM TO ODOO CHATTER ══
When user confirms to save the MOM:
1. Call odoo_search on project.task (fields=[id,name,project_id], domain=[]) to find the right task
2. Call odoo_post_chatter with model="project.task", record_id=<task_id>, body=<full MOM text>
3. Confirm: "✅ تم حفظ المحضر في Odoo على المهمة '[task name]' (ID=[id])"

══ CREATING ODOO TASKS FROM ACTION ITEMS ══
When user confirms to create tasks:
1. Call odoo_search on project.project (fields=[id,name], domain=[]) to find project_id
2. Call odoo_create on project.task with:
   - name: action item description
   - project_id: integer ID found in step 1
   - date_deadline: YYYY-MM-DD (from action item)
3. Confirm each: "✅ تم إنشاء المهمة '[name]' بـ ID=[id] في مشروع '[project]'"

For calendar.event creation (schedule a follow-up meeting):
- name: meeting title
- start: "YYYY-MM-DD 09:00:00"
- stop: "YYYY-MM-DD 10:00:00"

══ SCOPE ══
You operate within Odoo ERP only. Answer questions about business data, create Odoo records, and generate meeting documentation. Refuse unrelated questions politely.

RULES:
- ALWAYS search before creating — never guess or invent integer IDs
- Batch all needed searches in ONE response before creating
- Confirm every creation with its Odoo record ID
- Format output with markdown — tables for action items, headers for MOM sections
- Arabic or English — always match the user's language
- For MOM: use bilingual labels (Arabic / English) in all section headers

CHARTS: CHART_BAR:{{"title":"T","labels":["A","B"],"data":[10,20]}}  CHART_PIE:{{"title":"T","labels":["A"],"data":[1]}}"""

# ── Flask app ──────────────────────────────────────────────────────────────────
app = Flask(__name__)

# Support running under /ai path on the server
app.config["APPLICATION_ROOT"] = "/"

# ── Visitor tracking ───────────────────────────────────────────────────────────
# Set ADMIN_SECRET in Railway Variables to protect the /admin/visitors endpoint.
ADMIN_SECRET  = os.environ.get("ADMIN_SECRET", "").strip()
_VISITORS_DB  = os.environ.get("VISITORS_DB", "./visitors.db")   # set to /data/visitors.db with Railway Volume for full persistence
_visitors     = {}
_visit_lock   = threading.Lock()
_STATIC_EXTS  = {".js", ".css", ".png", ".ico", ".svg", ".woff", ".woff2", ".ttf", ".map"}

# ── Persistent storage: PostgreSQL (primary) or SQLite (fallback) ──────────────
# PostgreSQL survives Railway redeploys; SQLite is reset on every deploy.
# Add the Railway PostgreSQL plugin (free) → it auto-sets DATABASE_URL.
import sqlite3 as _sqlite3

_DATABASE_URL = os.environ.get("DATABASE_URL", "").strip()
_USE_PG       = False
_vdb          = None        # SQLite connection (only used when PG unavailable)
_vdb_lock     = threading.Lock()

_CREATE_TABLE_SQL = """CREATE TABLE IF NOT EXISTS visitors (
    ip TEXT PRIMARY KEY, device TEXT, browser TEXT, os TEXT, ua TEXT,
    first_seen TEXT, last_seen TEXT, requests INTEGER DEFAULT 1
)"""

def _pg_conn():
    """Open a fresh autocommit PostgreSQL connection. Caller must close it."""
    import psycopg2
    c = psycopg2.connect(_DATABASE_URL)
    c.autocommit = True
    return c

def _init_visitors_db():
    global _USE_PG, _vdb
    if _DATABASE_URL:
        try:
            conn = _pg_conn()
            with conn.cursor() as cur:
                cur.execute(_CREATE_TABLE_SQL)
            conn.close()
            _USE_PG = True
            logging.info("Visitors: PostgreSQL backend (data persists across redeploys)")
            return
        except Exception as _e:
            logging.warning("PostgreSQL unavailable (%s) — falling back to SQLite", _e)
    # SQLite fallback (resets on redeploy unless Railway Volume is mounted)
    try:
        _vdb = _sqlite3.connect(_VISITORS_DB, check_same_thread=False, timeout=10)
        _vdb.execute(_CREATE_TABLE_SQL)
        _vdb.commit()
        logging.info("Visitors: SQLite backend (%s)", _VISITORS_DB)
    except Exception as _e:
        logging.warning("Visitors DB init failed (%s) — in-memory only", _e)

_init_visitors_db()

def _db_upsert(ip, d):
    if _USE_PG:
        try:
            conn = _pg_conn()
            with conn.cursor() as cur:
                cur.execute(
                    """INSERT INTO visitors(ip,device,browser,os,ua,first_seen,last_seen,requests)
                       VALUES(%s,%s,%s,%s,%s,%s,%s,%s)
                       ON CONFLICT(ip) DO UPDATE
                       SET last_seen=EXCLUDED.last_seen, requests=EXCLUDED.requests""",
                    (ip, d["device"], d["browser"], d["os"], d["ua"],
                     d["first_seen"], d["last_seen"], d["requests"]))
            conn.close()
        except Exception as _e:
            logging.warning("PG upsert error: %s", _e)
    elif _vdb:
        with _vdb_lock:
            try:
                _vdb.execute(
                    """INSERT INTO visitors(ip,device,browser,os,ua,first_seen,last_seen,requests)
                       VALUES(?,?,?,?,?,?,?,?)
                       ON CONFLICT(ip) DO UPDATE SET last_seen=excluded.last_seen, requests=excluded.requests""",
                    (ip, d["device"], d["browser"], d["os"], d["ua"],
                     d["first_seen"], d["last_seen"], d["requests"]))
                _vdb.commit()
            except Exception as _e:
                logging.warning("SQLite write error: %s", _e)

def _db_load():
    if _USE_PG:
        try:
            conn = _pg_conn()
            with conn.cursor() as cur:
                cur.execute("SELECT ip,device,browser,os,ua,first_seen,last_seen,requests FROM visitors ORDER BY last_seen DESC")
                rows = cur.fetchall()
            conn.close()
            return {r[0]: {"device":r[1],"browser":r[2],"os":r[3],"ua":r[4],
                           "first_seen":r[5],"last_seen":r[6],"requests":r[7]} for r in rows}
        except Exception as _e:
            logging.warning("PG load error: %s", _e)
            return {}
    if _vdb:
        try:
            rows = _vdb.execute("SELECT ip,device,browser,os,ua,first_seen,last_seen,requests FROM visitors").fetchall()
            return {r[0]: {"device":r[1],"browser":r[2],"os":r[3],"ua":r[4],
                           "first_seen":r[5],"last_seen":r[6],"requests":r[7]} for r in rows}
        except Exception as _e:
            logging.warning("SQLite load error: %s", _e)
    return {}

# Load persisted data into memory at startup
_visitors.update(_db_load())
logging.info("Visitors loaded from DB: %d records (backend: %s)",
             len(_visitors), "PostgreSQL" if _USE_PG else "SQLite")

# ── Parser helpers ─────────────────────────────────────────────────────────────
def _parse_device(ua):
    u = ua or ""
    if any(k in u for k in ("iPhone","Android","Mobile","BlackBerry","Windows Phone","webOS")):
        return "📱 Mobile"
    if any(k in u for k in ("iPad","Tablet")):
        return "📟 Tablet"
    return "🖥 Desktop"

def _parse_browser(ua):
    u = ua or ""
    if "Edg/"    in u: return "Edge"
    if "OPR/"    in u or "Opera" in u: return "Opera"
    if "Chrome/" in u: return "Chrome"
    if "Firefox/" in u: return "Firefox"
    if "Safari/" in u: return "Safari"
    return "Other"

def _parse_os(ua):
    u = ua or ""
    if "Windows" in u: return "Windows"
    if "Mac OS"  in u: return "macOS"
    if "Android" in u: return "Android"
    if "iPhone"  in u or "iPad" in u: return "iOS"
    if "Linux"   in u: return "Linux"
    return "Other"

@app.before_request
def _track_visitor():
    path = request.path
    if any(path.endswith(ext) for ext in _STATIC_EXTS):
        return
    ip  = (request.headers.get("X-Forwarded-For") or request.remote_addr or "unknown").split(",")[0].strip()
    ua  = request.headers.get("User-Agent", "")[:200]
    now = datetime.datetime.utcnow().strftime("%Y-%m-%d %H:%M") + " UTC"
    with _visit_lock:
        if ip in _visitors:
            _visitors[ip]["last_seen"] = now
            _visitors[ip]["requests"] += 1
        else:
            if len(_visitors) >= 2000:
                oldest = min(_visitors, key=lambda k: _visitors[k]["first_seen"])
                del _visitors[oldest]
            _visitors[ip] = {
                "first_seen": now, "last_seen": now, "requests": 1,
                "ua": ua, "device": _parse_device(ua),
                "browser": _parse_browser(ua), "os": _parse_os(ua),
            }
        _db_upsert(ip, _visitors[ip])
    logging.info("VISIT ip=%s device=%s path=%s", ip, _parse_device(ua), path)

@app.route("/admin/visitors")
@app.route("/ai/admin/visitors")
def admin_visitors():
    # ── Secret check — only owner can see this ──────────────────────
    if ADMIN_SECRET:
        provided = (request.args.get("secret") or "").strip()
        if provided != ADMIN_SECRET:
            return jsonify({"error": "Unauthorized — wrong secret"}), 403
    with _visit_lock:
        snap = dict(_visitors)
    rows = sorted(snap.items(), key=lambda x: x[1]["last_seen"], reverse=True)
    visitors_out = [{"ip": ip, **v} for ip, v in rows]
    devices  = {}
    browsers = {}
    for v in snap.values():
        devices[v["device"]]   = devices.get(v["device"],   0) + 1
        browsers[v["browser"]] = browsers.get(v["browser"], 0) + 1
    return jsonify({
        "unique_ips":     len(snap),
        "total_requests": sum(v["requests"] for v in snap.values()),
        "devices":  devices,
        "browsers": browsers,
        "visitors": visitors_out,
    })

@app.route("/")
@app.route("/ai")
@app.route("/ai/")
def index():
    return render_template("index.html")

@app.route("/chat", methods=["POST"])
@app.route("/ai/chat", methods=["POST"])
def chat():
    data         = request.json
    history      = data.get("messages", [])
    override_key = data.get("api_key", "").strip() or None
    mode         = data.get("mode", "assistant")     # "assistant" or "implementer"

    _providers = _prov_mgr.providers(override_key)
    if not _providers:
        return jsonify({"error": "No AI provider configured. Set GEMINI_API_KEY or GROQ_API_KEY in Railway Variables."}), 400

    _system_prompt = get_implementer_system_prompt() if mode == "implementer" else get_system_prompt()
    _tools         = IMPLEMENTER_TOOLS               if mode == "implementer" else TOOLS
    _tool_runner   = _run_implementer_tool            if mode == "implementer" else run_tool

    def generate():
        _req_id = f"req_{int(time.time()*1000) % 1000000}"
        try:
            # ── Build base messages ─────────────────────────────────────────
            recent    = history[-4:] if len(history) > 4 else history
            base_msgs = [{"role": "system", "content": _system_prompt}]
            for m in recent:
                base_msgs.append({"role": m["role"], "content": m.get("content") or ""})

            # ── Threaded API call helper ────────────────────────────────────
            # Generator: yields SSE pings while the call runs; returns (resp, err).
            # Caller uses: resp, err = yield from _ai_call(cl, model, msgs, tools)
            def _ai_call(cl, model, msgs, tools=None):
                _r = [None]; _e = [None]; _ev = threading.Event()
                def _worker():
                    try:
                        kw = dict(model=model, messages=msgs,
                                  max_tokens=1800, temperature=0.1)
                        if tools:
                            kw["tools"]       = tools
                            kw["tool_choice"] = "auto"
                        # Auto-fix unsupported params (gpt-5 series quirks):
                        # 1st 400 → max_completion_tokens; 2nd → drop temperature
                        _last_exc = None
                        for _ in range(3):
                            try:
                                _r[0] = cl.chat.completions.create(**kw)
                                break
                            except Exception as _exc:
                                _em = str(_exc); _last_exc = _exc
                                if "max_tokens" in _em and "max_completion_tokens" in _em:
                                    kw.pop("max_tokens", None)
                                    kw["max_completion_tokens"] = 1800
                                elif "temperature" in _em:
                                    kw.pop("temperature", None)
                                else:
                                    _e[0] = _exc; break
                        else:
                            _e[0] = _last_exc
                    except Exception as exc:
                        _e[0] = exc
                    finally:
                        _ev.set()
                threading.Thread(target=_worker, daemon=True).start()
                while not _ev.wait(timeout=20):
                    yield f"data: {json.dumps({'type': 'ping'})}\n\n"
                return _r[0], _e[0]

            # ── State ───────────────────────────────────────────────────────
            _prov_idx         = 0     # current index into _providers list
            _tool_call_counts = {}    # call_key → count (cross-round loop detection)
            _tool_chunks      = []    # accumulated raw tool results for Phase 2 synthesis
            _xml_guard_count  = 0
            _answered         = False
            messages          = list(base_msgs)

            # ── Phase 1: Tool-calling loop (max 4 rounds) ───────────────────
            for _round in range(4):
                if _prov_idx >= len(_providers):
                    yield f"data: {json.dumps({'type': 'text', 'text': _prov_mgr.user_message('no_providers')})}\n\n"
                    _answered = True
                    break

                _p     = _providers[_prov_idx]
                _cl    = _p["client"]
                _model = _p["model"]
                logging.warning("[%s] Phase1 round=%d prov=%s model=%s",
                                _req_id, _round, _p["name"], _model)

                resp, err = yield from _ai_call(_cl, _model, messages, _tools)

                if err is not None:
                    cat, can_fallback = _prov_mgr.classify_error(err)
                    logging.warning("[%s] Phase1 err cat=%s fallback=%s: %s",
                                    _req_id, cat, can_fallback, str(err)[:200])
                    if can_fallback and _prov_idx + 1 < len(_providers):
                        _prov_idx += 1
                        _np = _providers[_prov_idx]
                        _sw_msg = f"{_p['name']} error — trying {_np['name']} ({_np['model']})…"
                        yield f"data: {json.dumps({'type': 'tool', 'name': 'switch_provider', 'input': {'model': _sw_msg}})}\n\n"
                        messages = list(base_msgs)  # clean slate for new provider
                        continue
                    yield f"data: {json.dumps({'type': 'text', 'text': _prov_mgr.user_message(cat)})}\n\n"
                    _answered = True
                    break

                msg    = resp.choices[0].message
                finish = resp.choices[0].finish_reason
                logging.warning("[%s] Phase1 round=%d finish=%s tools=%d content=%d",
                                _req_id, _round, finish,
                                len(msg.tool_calls or []), len(msg.content or ""))

                # Direct text answer — no tools needed
                if not msg.tool_calls:
                    text = msg.content or ""

                    # Guard: model leaked <function= XML syntax into content
                    if "<function=" in text:
                        _xml_guard_count += 1
                        logging.warning("[%s] XML leak round=%d count=%d",
                                        _req_id, _round, _xml_guard_count)
                        if _xml_guard_count <= 1:
                            messages.append({"role": "assistant", "content": text})
                            messages.append({"role": "user", "content":
                                "IMPORTANT: Do NOT output function call syntax. "
                                "Write only the final human-readable answer in the user's language."})
                            continue
                        text = re.sub(r'<function[^>]*>[\s\S]*?</function[^>]*>', '', text).strip()
                        text = re.sub(r'</?function[^>]*>', '', text).strip()
                        if not text:
                            text = _prov_mgr.user_message("bad_request")

                    # Guard: empty first response — prompt for tool use
                    if not text and _round == 0 and not _tool_chunks:
                        logging.warning("[%s] Empty response round 0, prompting tool use", _req_id)
                        messages.append({"role": "assistant", "content": ""})
                        messages.append({"role": "user", "content":
                            "Call the appropriate Odoo tool now to retrieve the data. "
                            "Do not answer from memory."})
                        continue

                    if not text:
                        text = _prov_mgr.user_message("unknown_error")
                    yield f"data: {json.dumps({'type': 'text', 'text': text})}\n\n"
                    _answered = True
                    break

                # Build assistant message with tool calls
                messages.append({
                    "role": "assistant",
                    "content": msg.content or "",
                    "tool_calls": [
                        {"id": tc.id, "type": "function",
                         "function": {"name": tc.function.name,
                                      "arguments": tc.function.arguments}}
                        for tc in msg.tool_calls
                    ]
                })

                # Dedup within batch
                _seen_keys = set()
                _batch     = []
                for tc in msg.tool_calls:
                    try:    args = json.loads(tc.function.arguments)
                    except: args = {}
                    _k = f"{tc.function.name}:{args.get('model','') if isinstance(args, dict) else ''}"
                    _batch.append((tc, args, _k, _k in _seen_keys))
                    _seen_keys.add(_k)

                # Cross-round loop: same key seen in a prior round
                _loop = any(_tool_call_counts.get(k, 0) >= 1
                            for _, _, k, d in _batch if not d)

                # Notify client + count unique calls
                for tc, args, k, is_dup in _batch:
                    if not is_dup:
                        yield f"data: {json.dumps({'type': 'tool', 'name': tc.function.name, 'input': args})}\n\n"
                        _tool_call_counts[k] = _tool_call_counts.get(k, 0) + 1

                # Execute all non-dup tools in parallel
                _tres   = {}
                _t_lock = threading.Lock()
                _pend   = [sum(1 for _, _, _, d in _batch if not d)]
                _all    = threading.Event()
                if _pend[0] == 0:
                    _all.set()

                def _do_tool(name, a, tid, _tres=_tres, _t_lock=_t_lock,
                             _pend=_pend, _all=_all):
                    try:   res = _tool_runner(name, a)
                    except Exception as _te: res = json.dumps({"error": str(_te)})
                    with _t_lock:
                        _tres[tid] = res
                        _pend[0] -= 1
                        if _pend[0] == 0:
                            _all.set()

                for tc, args, k, is_dup in _batch:
                    if not is_dup:
                        threading.Thread(target=_do_tool,
                                         args=(tc.function.name, args, tc.id),
                                         daemon=True).start()

                # Wait up to 45 s; ping every 5 s
                _w = 0
                while _w < 45 and not _all.is_set():
                    _all.wait(timeout=5); _w += 5
                    if not _all.is_set() and _w < 45:
                        yield f"data: {json.dumps({'type': 'ping'})}\n\n"

                # Append tool results to messages + accumulate for synthesis
                for tc, args, k, is_dup in _batch:
                    if is_dup:
                        c = json.dumps({"note": "duplicate query skipped — same result applies"})
                    else:
                        c = _tres.get(tc.id)
                        if not c:
                            c = json.dumps({"error": "Tool timed out after 45 s."})
                            logging.warning("[%s] Tool %s timed out", _req_id, tc.function.name)
                    messages.append({"role": "tool", "tool_call_id": tc.id, "content": c})
                    _tool_chunks.append(c)

                # If loop detected, skip next round and go straight to Phase 2
                if _loop:
                    logging.warning("[%s] Tool loop detected round=%d — forcing Phase 2",
                                    _req_id, _round)
                    break

                # Inject synthesis hint; if AI responds with text on next round → done
                messages.append({
                    "role": "user",
                    "content": (
                        "The Odoo data above has been retrieved. "
                        "Give the final answer now in the same language as the original question. "
                        "Use markdown tables. Do NOT output any function call syntax."
                    )
                })
                # Continue → model may answer directly next round or call more tools

            # ── Phase 2: Synthesis with clean history ───────────────────────
            # Runs only when Phase 1 collected tool results but didn't produce a text answer.
            if not _answered and _tool_chunks:
                # Strip tool-call metadata: Gemini re-infers tool schemas from history,
                # so sending clean messages is the only reliable way to stop it looping.
                _clean = []
                for _m in messages:
                    _mr = _m.get("role"); _mc = _m.get("content") or ""
                    if _mr == "tool":
                        pass  # collected in _tool_chunks
                    elif _mr == "assistant" and _m.get("tool_calls"):
                        pass  # omit — carries implicit schema
                    elif _mr == "user" and (
                        _mc.startswith("The Odoo data above") or
                        _mc.startswith("Call the appropriate Odoo tool")
                    ):
                        pass  # omit injected prompts
                    else:
                        _clean.append(_m)

                _clean.append({
                    "role": "user",
                    "content": (
                        "Odoo data retrieved:\n\n" +
                        "\n\n---\n\n".join(_tool_chunks) +
                        "\n\nAnswer the original question above. "
                        "Use markdown tables. Reply in the same language as the question. "
                        "Do NOT output any function call syntax or tool invocations — just the answer."
                    )
                })

                # Try all providers for synthesis — start from 0 (fresh attempt)
                for _syn_idx in range(len(_providers)):
                    _sp    = _providers[_syn_idx]
                    logging.warning("[%s] Phase2 attempt=%d prov=%s model=%s msgs=%d",
                                    _req_id, _syn_idx, _sp["name"], _sp["model"], len(_clean))

                    resp2, err2 = yield from _ai_call(_sp["client"], _sp["model"], _clean)

                    if err2 is not None:
                        cat2, can_fb2 = _prov_mgr.classify_error(err2)
                        logging.warning("[%s] Phase2 err cat=%s: %s",
                                        _req_id, cat2, str(err2)[:200])
                        if can_fb2 and _syn_idx + 1 < len(_providers):
                            _np2 = _providers[_syn_idx + 1]
                            _sw2_msg = f"Synthesis failed on {_sp['name']} — trying {_np2['name']}…"
                            yield f"data: {json.dumps({'type': 'tool', 'name': 'switch_provider', 'input': {'model': _sw2_msg}})}\n\n"
                            continue
                        yield f"data: {json.dumps({'type': 'text', 'text': _prov_mgr.user_message(cat2)})}\n\n"
                        _answered = True
                        break

                    text2 = resp2.choices[0].message.content or ""
                    if "<function=" in text2:
                        text2 = re.sub(r'<function[^>]*>[\s\S]*?</function[^>]*>', '', text2).strip()
                        text2 = re.sub(r'</?function[^>]*>', '', text2).strip()
                    if not text2:
                        text2 = _prov_mgr.user_message("unknown_error")
                    yield f"data: {json.dumps({'type': 'text', 'text': text2})}\n\n"
                    _answered = True
                    break

            if not _answered:
                yield f"data: {json.dumps({'type': 'text', 'text': '⚠️ Could not complete the request. Please try again.'})}\n\n"

        except Exception as outer_err:
            # Last-resort catch — always send something to unblock the UI
            yield f"data: {json.dumps({'type': 'text', 'text': f'❌ Unexpected error: {str(outer_err)[:300]}'})}\n\n"

        finally:
            # Always send done so the frontend never hangs
            yield f"data: {json.dumps({'type': 'done'})}\n\n"

    return Response(stream_with_context(generate()),
                    mimetype="text/event-stream",
                    headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})

@app.route("/ai/health")
@app.route("/health")
def health_check():
    """Startup / liveness health check — tests both AI providers with a minimal call."""
    import concurrent.futures, datetime

    def _probe(p):
        name  = p["name"]
        model = p["model"]
        t0    = datetime.datetime.utcnow()
        try:
            _probe_kw = dict(
                model=model,
                messages=[{"role": "user", "content": "Reply with the single word OK."}],
                max_tokens=5,
                temperature=0,
            )
            _plast = None
            for _ in range(3):
                try:
                    p["client"].chat.completions.create(**_probe_kw); break
                except Exception as _fe:
                    _pem = str(_fe); _plast = _fe
                    if "max_tokens" in _pem and "max_completion_tokens" in _pem:
                        _probe_kw.pop("max_tokens", None)
                        _probe_kw["max_completion_tokens"] = 5
                    elif "temperature" in _pem:
                        _probe_kw.pop("temperature", None)
                    else:
                        raise
            else:
                raise _plast
            ms = int((datetime.datetime.utcnow() - t0).total_seconds() * 1000)
            return {"provider": name, "model": model, "status": "ok", "latency_ms": ms}
        except Exception as exc:
            ms  = int((datetime.datetime.utcnow() - t0).total_seconds() * 1000)
            cat, _ = _prov_mgr.classify_error(exc)
            return {"provider": name, "model": model, "status": "error",
                    "error_category": cat, "detail": str(exc)[:200], "latency_ms": ms}

    providers = _prov_mgr.providers()
    if not providers:
        return jsonify({"status": "unhealthy", "reason": "no_providers_configured",
                        "providers": []}), 503

    results = []
    with concurrent.futures.ThreadPoolExecutor(max_workers=2) as ex:
        futures = {ex.submit(_probe, p): p for p in providers}
        for fut in concurrent.futures.as_completed(futures):
            results.append(fut.result())

    ok_count = sum(1 for r in results if r["status"] == "ok")
    overall  = "healthy" if ok_count > 0 else "unhealthy"
    http_code = 200 if ok_count > 0 else 503

    return jsonify({
        "status":    overall,
        "providers": results,
        "primary_ok": any(r["status"] == "ok" and r["provider"] == "Gemini" for r in results),
        "fallback_ok": any(r["status"] == "ok" and r["provider"] == "Groq" for r in results),
    }), http_code

@app.route("/reports-data")
@app.route("/ai/reports-data")
def reports_data():
    def safe(model, method, args, kwargs=None):
        try:
            return odoo_call(model, method, args, kwargs or {})
        except Exception as e:
            return {"error": str(e)}

    # Construction: BOQ lines by project
    boq_by_project = safe("project.subcontracting.boq.line", "read_group",
        [[], ["project_id", "boq_cost:sum", "quantity:sum"], ["project_id"]], {"lazy": False})
    # Construction: subcontractor contracts by project
    boq_contracts_by_project = safe("subcontractor.contract", "read_group",
        [[], ["project_id", "contract_value:sum", "bills_amount_total:sum"], ["project_id"]], {"lazy": False})
    # Construction: advance payments by state
    adv_by_state = safe("construction.advance.payment", "read_group",
        [[], ["state"], ["state"]], {"lazy": False})
    # Construction: advance payments by project
    adv_by_project = safe("construction.advance.payment", "read_group",
        [[], ["project_id", "amount:sum", "due_amount:sum"], ["project_id"]], {"lazy": False})
    # Construction: detailed items progress by project
    items_by_project = safe("project.detailed.item.line", "read_group",
        [[], ["project_id", "total_cost:sum", "actual_cost:sum"], ["project_id"]], {"lazy": False})
    # Purchases: top 10 vendors by total
    po_by_vendor = safe("purchase.order", "read_group",
        [[["state", "in", ["purchase", "done"]]], ["partner_id", "amount_total:sum"], ["partner_id"]],
        {"lazy": False, "limit": 10, "orderby": "amount_total desc"})
    # HR: employees by department
    emp_by_dept = safe("hr.employee", "read_group",
        [[], ["department_id"], ["department_id"]], {"lazy": False})
    # Real Estate: units by usability_type
    units_by_type = safe("rs.dev.unit", "read_group",
        [[], ["usability_type"], ["usability_type"]], {"lazy": False})
    # Construction: detailed BOQ lines by project (client-facing BOQ) — total_cost is the correct amount
    boq_detail_by_project = safe("project.detailed.item.line", "read_group",
        [[], ["project_id", "total_cost:sum"], ["project_id"]], {"lazy": False})
    # Construction: plan milestones by plan_type
    plan_by_type = safe("main.project.plan.operation", "read_group",
        [[], ["plan_type"], ["plan_type"]], {"lazy": False})
    # Real Estate: units by historical_process (correct status field on rs.dev.unit)
    units_by_hp = safe("rs.dev.unit", "read_group",
        [[], ["historical_process"], ["historical_process"]], {"lazy": False})

    # Inventory: stock on hand by location (internal locations only, qty > 0)
    inv_by_location = safe("stock.quant", "read_group",
        [[["location_id.usage","=","internal"],["quantity",">",0]],
         ["location_id","quantity:sum"], ["location_id"]],
        {"lazy": False, "limit": 10, "orderby": "quantity desc"})
    # Inventory: active transfers by state (exclude cancelled/draft)
    inv_by_state = safe("stock.picking", "read_group",
        [[["state","not in",["cancel","draft"]]], ["state"], ["state"]], {"lazy": False})
    # Inventory: transfers by operation type (receipts vs deliveries vs internal)
    inv_by_type = safe("stock.picking", "read_group",
        [[["state","not in",["cancel"]]], ["picking_type_id"], ["picking_type_id"]],
        {"lazy": False})

    def extract(rows, label_field, count_field="__count", amount_field=None):
        if isinstance(rows, dict) and "error" in rows:
            return {"error": rows["error"]}
        out = []
        for r in rows:
            lbl = r.get(label_field)
            if isinstance(lbl, (list, tuple)):
                lbl = lbl[1]
            elif not lbl:
                lbl = "غير محدد"
            entry = {"label": str(lbl), "count": r.get(count_field, 0)}
            if amount_field:
                entry["amount"] = r.get(amount_field, 0)
            out.append(entry)
        return out

    return jsonify({
        "boq_by_project":           extract(boq_by_project,           "project_id",        amount_field="boq_cost"),
        "boq_contracts_by_project": extract(boq_contracts_by_project, "project_id",        amount_field="contract_value"),
        "boq_detail_by_project":    extract(boq_detail_by_project,    "project_id",        amount_field="total_cost"),
        "adv_by_state":             extract(adv_by_state,             "state"),
        "adv_by_project":           extract(adv_by_project,           "project_id",        amount_field="amount"),
        "items_by_project":         extract(items_by_project,         "project_id",        amount_field="total_cost"),
        "po_by_vendor":             extract(po_by_vendor,             "partner_id",        amount_field="amount_total"),
        "emp_by_dept":              extract(emp_by_dept,              "department_id"),
        "units_by_state":           extract(units_by_hp,              "historical_process"),
        "units_by_type":            extract(units_by_type,            "usability_type"),
        "plan_by_state":            extract(plan_by_type,             "plan_type"),
        "inv_by_location":          extract(inv_by_location,          "location_id",   amount_field="quantity"),
        "inv_by_state":             extract(inv_by_state,             "state"),
        "inv_by_type":              extract(inv_by_type,              "picking_type_id"),
    })

# ── BOQ Import ─────────────────────────────────────────────────────────────────

@app.route("/upload-boq")
@app.route("/ai/upload-boq")
def upload_boq_page():
    return render_template("upload_boq.html")

@app.route("/parse-boq", methods=["POST"])
@app.route("/ai/parse-boq", methods=["POST"])
def parse_boq():
    """Read uploaded Excel, extract rows, use AI to suggest column mapping."""
    try:
        return _parse_boq_impl()
    except Exception as e:
        logging.exception("Unhandled error in parse_boq")
        return jsonify({"error": f"Server error: {e}"}), 500

import unicodedata as _ucd

def _norm_header(h):
    """Normalize Arabic header: alef variants, alef-maqsura to ya, ta-marbuta, harakat, invisible chars."""
    h = _ucd.normalize("NFKC", str(h))
    cleaned = []
    for ch in h:
        cp = ord(ch)
        if (0x200B <= cp <= 0x200F or 0x202A <= cp <= 0x202E or
                0x2060 <= cp <= 0x2069 or cp in (0xFEFF, 0xA0)):
            cleaned.append(" ")
        else:
            cleaned.append(ch)
    h = "".join(cleaned)
    h = re.sub("[\u0623\u0625\u0622\u0671]", "\u0627", h)  # alef variants -> bare alef
    h = h.replace("\u0649", "\u064A")  # alef maqsura (\u0649) -> ya (\u064A)
    h = h.replace("\u0629", "\u0647")  # ta marbuta -> ha
    h = re.sub("[\u064B-\u065F]", "", h)  # strip harakat
    return h.strip().lower()

_BOQ_KEYWORD_RULES = [
    # item_code MUST come before name so "item code"/"item no" headers are claimed first,
    # preventing "item" (added to name keywords below) from matching them.
    ("item_code",         ["القسم", "قسم", "رقم البند",
                            "item code", "item no", "task code", "line code", "line no", "serial"]),
    # name comes second; "item" / "items" / "item description" are common English column names
    ("name",              ["البند", "بيان", "الوصف", "وصف",
                            "description", "task code description", "work description",
                            "item name", "item description", "activity",
                            "item", "items"]),
    ("quantity",          ["الكمية", "كمية", "كميه", "quantity", "qty"]),
    ("unit",              ["الوحدة", "وحدة", "وحده",
                            "description uom", "uom", "unit of measure", "task code unit", "unit"]),
    ("unit_cost",         ["الفئة", "فئة", "سعر الوحدة", "سعر",
                            "unit cost", "unit price", "unit rate", "rate", "boq rate"]),
    ("total_cost",        ["الاجمالي", "اجمالي", "المبلغ",
                            "total", "amount", "subtotal", "total cost", "total amount", "boq amount"]),
    ("main_item_line_id", ["main item line id"]),
    ("section_name",      ["chapter name", "section header", "chapter title", "main chapter"]),
    ("boq_item_id",       ["description id", "boq item id"]),
    ("work_type",         ["نوع العمل", "description categ", "categ", "work type", "category",
                            "trade", "division", "discipline"]),
    ("notes",             ["ملاحظات", "resource description", "notes", "remarks"]),
]
_BOQ_SKIP_KEYWORDS = ["#n/a", "cost code", "wastage", "usage", "resource code",
                       "billed", "remain", "col0"]

# Keywords that flag a sheet as a resource-analysis sheet (not a BOQ sheet)
_RESOURCE_SHEET_PENALTIES = ["resource code", "wastage", "usage", "b.o.m", "bom", "final rate"]

def _boq_sheet_score(header_row, row_count):
    """Score a worksheet for BOQ suitability — higher = more likely the main BOQ sheet."""
    score = 0
    non_null = [h for h in header_row if h is not None and str(h).strip()]
    if len(non_null) < 2:
        return -999
    for h in header_row:
        if not h:
            continue
        hl = _norm_header(str(h))
        # Heavy penalty for resource-analysis column headers
        if any(kw in hl for kw in _RESOURCE_SHEET_PENALTIES):
            score -= 30
            continue
        # Reward matching any known BOQ field keyword
        for _field, keywords in _BOQ_KEYWORD_RULES:
            if any(_norm_header(kw) in hl for kw in keywords):
                score += 15
                break
        else:
            # Small bonus for Arabic content even if no keyword match
            if any('؀' <= c <= 'ۿ' for c in str(h)):
                score += 2
    # Penalise extremely large sheets (resource breakdowns can have 5 000+ rows)
    if row_count and row_count > 1200:
        score -= 20
    return score

def _keyword_map_boq(headers):
    """Map column indices to BOQ field names using Arabic/English keywords."""
    result = {}
    claimed_fields = set()
    for i, h in enumerate(headers):
        hl = _norm_header(h)
        if not hl or any(kw in hl for kw in _BOQ_SKIP_KEYWORDS):
            continue
        # Skip auto-generated no-header names like Col0, Col1, Col2 ...
        if re.match(r'^col\d+$', hl):
            continue
        for field, keywords in _BOQ_KEYWORD_RULES:
            if field in claimed_fields:
                continue
            if any(_norm_header(kw) in hl for kw in keywords):
                result[str(i)] = field
                claimed_fields.add(field)
                break
    logging.info("Keyword BOQ map: %s", result)
    return result
def _parse_boq_impl():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    f = request.files["file"]
    _boq_override = request.form.get("api_key", "").strip() or None

    try:
        import openpyxl
        raw_bytes = f.read()
        # Single open — use max_row metadata (instant, no row iteration) to pick sheet
        wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
        best_ws   = wb.active.title if wb.active else (wb.sheetnames[0] if wb.sheetnames else None)
        best_score = -999

        for sname in wb.sheetnames:
            ws_try    = wb[sname]
            row_count = ws_try.max_row or 0
            if row_count < 2:
                continue
            # Read first ≤10 rows to find the header and score the sheet
            header_row = None
            sampled = 0
            for row_vals in ws_try.iter_rows(min_row=1, max_row=10, values_only=True):
                sampled += 1
                non_empty = [c for c in row_vals if c is not None and str(c).strip()]
                if len(non_empty) >= 3:
                    header_row = row_vals
                    break
            if header_row is None:
                continue
            score = _boq_sheet_score(header_row, row_count)
            logging.info("Sheet '%s': score=%d, rows=%d", sname, score, row_count)
            if score > best_score:
                best_score, best_ws = score, sname

        ws = wb[best_ws] if best_ws else wb.active
        if ws is None:
            wb.close()
            return jsonify({"error": "Could not read any sheet from the Excel file"}), 400
        logging.info("Selected sheet '%s' (score=%d)", best_ws, best_score)
        # Read up to 600 rows (header search + up to 300 data rows with buffer)
        raw_rows = []
        for row in ws.iter_rows(values_only=True):
            raw_rows.append(row)
            if len(raw_rows) >= 600:
                break
        wb.close()
    except Exception as e:
        return jsonify({"error": f"Failed to read Excel file: {e}"}), 400

    if len(raw_rows) < 2:
        return jsonify({"error": "Excel file is empty or has no data"}), 400

    # Find first row that has at least 3 non-empty cells → use as headers
    headers = []
    data_start = 0
    for i, row in enumerate(raw_rows):
        non_empty = [c for c in row if c is not None and str(c).strip()]
        if len(non_empty) >= 3:
            headers = [str(c).strip() if c is not None else f"Col{j}" for j, c in enumerate(row)]
            data_start = i + 1
            break

    if not headers:
        return jsonify({"error": "Could not detect a header row (need ≥3 non-empty cells)"}), 400

    # Collect up to 300 data rows (skip fully empty rows)
    data_rows = []
    for row in raw_rows[data_start: data_start + 300]:
        values = [str(c).strip() if c is not None else "" for c in row]
        if any(v for v in values):
            data_rows.append(values)

    if not data_rows:
        return jsonify({"error": "No data rows found after the header"}), 400

    # Step 1: keyword-based mapping (works offline, handles Arabic + English)
    column_map = _keyword_map_boq(headers)

    # Step 2: AI mapping (enhances/overrides keyword map if AI is available)
    sample = data_rows[:3]
    prompt = (
        "You are mapping an Excel BOQ (Bill of Quantities) file to Odoo construction models.\n\n"
        "EXCEL COLUMNS (index: header | sample values):\n"
        + "\n".join(
            f"  {i}: {h} | {' / '.join(str(r[i]) for r in sample if i < len(r) and r[i])}"
            for i, h in enumerate(headers)
        )
        + "\n\nAVAILABLE ODOO FIELDS (map each column to exactly one):\n"
        "  COMMON (all BOQ models):\n"
        "    name          = البند / Description / item description (REQUIRED — must map this)\n"
        "    quantity      = الكمية / Quantity / Qty\n"
        "    unit_cost     = الفئة / Unit Cost / Rate / Unit Price\n"
        "    total_cost    = الإجمالي / Total / Amount\n"
        "    item_code     = القسم / رقم البند / Line Code / Item No\n"
        "    unit          = الوحدة / UOM / Unit of Measure / Description UOM\n"
        "    work_type     = نوع العمل / Category / Description Categ\n"
        "    notes         = ملاحظات / Remarks / Notes\n"
        "  project.main.item.line (chapter/section headers with no qty):\n"
        "    section_name  = Chapter name / Main Item Line (text, not ID)\n"
        "  project.detailed.item.line:\n"
        "    main_item_line_id = Main Item Line ID (INTEGER — Odoo record ID e.g. 537)\n"
        "    boq_item_id       = Description ID / BOQ Item ID (INTEGER)\n"
        "  project.subcontracting.boq.line:\n"
        "    boq_contract_id   = BOQ Contract ID (INTEGER)\n\n"
        "MAPPING RULES:\n"
        "  - البند/Description (not ID) → name\n"
        "  - الكمية/Quantity/Qty → quantity\n"
        "  - الوحدة/UOM/Description UOM → unit\n"
        "  - الفئة/Rate/Unit Cost/Unit Price → unit_cost\n"
        "  - الإجمالي/Total/Amount → total_cost\n"
        "  - القسم/Line Code/Item No → item_code\n"
        "  - 'Main Item Line ID' column with integers (537,538...) → main_item_line_id\n"
        "  - 'Main Item Line' column with text only → section_name\n"
        "  - 'Description ID' / 'Description Categ' → boq_item_id / work_type\n"
        "  - 'ITEM' / 'ITEMS' / 'ITEM DESCRIPTION' standalone column with text → name (NOT section_name)\n"
        "  - 'DIVISION' / 'TRADE' / 'DISCIPLINE' → work_type\n"
        "  - 'BOQ RATE' / 'BOQ UNIT RATE' → unit_cost\n"
        "  - 'BOQ AMOUNT' / 'BOQ TOTAL' → total_cost\n"
        "  - 'BOQ UNIT' / 'BOQ UOM' → unit\n"
        "  - Columns named 'Col0', 'Col1', 'Col2' (no header text) → SKIP or item_code for first one\n"
        "  - SKIP: #N/A columns, Cost Code, Wastage, Usage, Resource Code, billed, remain\n\n"
        "CRITICAL RULE: section_name is ONLY for chapter/section HEADERS that have NO quantity and NO price.\n"
        "If a column has descriptions AND quantities exist in the same rows → map it to name.\n\n"
        "Return ONLY a JSON object: {\"col_index\": \"field_name\", ...}\n"
        "Example: {\"1\":\"item_code\",\"2\":\"name\",\"3\":\"quantity\",\"4\":\"unit\",\"5\":\"unit_cost\",\"6\":\"total_cost\"}"
    )
    try:
        _boq_client, _boq_model = _make_chat_client(_boq_override)
        if _boq_client is not None:
            resp = _boq_client.chat.completions.create(
                model=_boq_model,
                messages=[{"role": "user", "content": prompt}],
                max_tokens=300, temperature=0,
                timeout=25,
            )
            raw_json = resp.choices[0].message.content.strip()
            # extract JSON — allow multiline
            m = re.search(r'\{[\s\S]*?\}', raw_json)
            if m:
                ai_map = json.loads(m.group())
                # AI map wins over keyword map
                column_map.update(ai_map)
                logging.info("AI column map: %s", ai_map)
    except Exception as e:
        logging.warning("AI column mapping failed (using keyword map): %s", e)

    return jsonify({
        "headers": headers,
        "rows": data_rows,
        "column_map": column_map,
        "total_rows": len(data_rows),
    })


@app.route("/get-projects", methods=["GET"])
@app.route("/ai/get-projects", methods=["GET"])
def get_projects():
    """Return list of projects from Odoo for the project selector."""
    try:
        projects = odoo_call("project.project", "search_read", [[]], {"fields": ["id", "name"], "limit": 100, "order": "name asc"})
        return jsonify(projects)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/import-boq", methods=["POST"])
@app.route("/ai/import-boq", methods=["POST"])
def import_boq():
    """Import mapped BOQ rows into Odoo models."""
    data       = request.json
    rows       = data.get("rows", [])
    column_map = data.get("column_map", {})   # {"col_index_str": "field_name"}
    project_id = data.get("project_id")
    models_to_import = data.get("models", ["project.subcontracting.boq.line", "project.detailed.item.line"])

    if not rows:
        return jsonify({"error": "No rows to import"}), 400
    if not column_map:
        return jsonify({"error": "Column mapping is required"}), 400
    if not project_id:
        return jsonify({"error": "Project is required"}), 400

    col_map = {int(k): v for k, v in column_map.items()}

    def _to_float(s):
        try:
            return float(str(s).replace(",", "").replace(" ", "") or 0)
        except Exception:
            return 0.0

    def _is_section_header(vals, qty, unit_cost):
        """True when a row is a chapter/section header (no numeric values)."""
        if qty != 0 or unit_cost != 0:
            return False
        code = str(vals.get("item_code", "")).strip()
        # Section codes look like "1", "2", "A", "I" — no dash or dot
        if code and '-' not in code and '.' not in code:
            return True
        # Or: no code at all but name is short (chapter title)
        if not code and vals.get("name") and len(vals["name"]) < 80:
            return True
        return False

    imported        = 0
    sections        = 0
    skipped         = 0
    errors          = []
    boq_contract_id = None
    current_main_id = None   # project.main.item.line id of current section
    last_name       = None   # carry-forward for merged-cell continuation rows
    last_item_code  = None

    def _err(e):
        s = getattr(e, "faultString", None) or str(e)
        return s.replace("<", "[").replace(">", "]")[:500]

    def _to_int_id(v):
        """Convert a cell value like '537' or '537.0' to int, or None."""
        try:
            return int(float(str(v).replace(",", "").strip()))
        except Exception:
            return None

    # ── Find existing boq.contract for the project (needed for subcontracting) ──
    if "project.subcontracting.boq.line" in models_to_import:
        try:
            existing_contracts = odoo_call("boq.contract", "search_read",
                [[["project_id", "=", project_id]]],
                {"fields": ["id", "name"], "limit": 1, "order": "id desc"})
            if existing_contracts:
                boq_contract_id = existing_contracts[0]["id"]
                logging.info("Using existing boq.contract id=%d '%s'",
                             boq_contract_id, existing_contracts[0]["name"])
            else:
                # Try creating a minimal contract
                from datetime import date as _date
                try:
                    boq_contract_id = odoo_call("boq.contract", "create",
                        [{"name": f"Tender Import {_date.today()}", "project_id": project_id}])
                except Exception as ce:
                    logging.warning("boq.contract create failed: %s — subcontracting lines will be skipped", ce)
                    models_to_import = [m for m in models_to_import
                                        if m != "project.subcontracting.boq.line"]
        except Exception as e:
            logging.warning("boq.contract lookup failed: %s", e)

    sub_ok = det_ok = 0

    for i, row in enumerate(rows):
        vals = {}
        for col_idx, field in col_map.items():
            if col_idx < len(row) and row[col_idx]:
                vals[field] = row[col_idx]

        # Pre-compute numeric values (needed for section vs item detection below)
        qty_pre       = _to_float(vals.get("quantity", 0))
        uc_pre        = _to_float(vals.get("unit_cost") or vals.get("unit_price", 0))
        tc_pre        = _to_float(vals.get("total_cost", 0))

        # ── section_name check FIRST ────────────────────────────────────────────
        # This must come before the name-skip so rows that only have section_name
        # are not silently dropped.
        explicit_sec = str(vals.get("section_name", "")).strip()
        if explicit_sec:
            if qty_pre or uc_pre or tc_pre:
                # Has quantities → section_name column was mis-mapped; treat as item name
                if not vals.get("name"):
                    vals["name"] = explicit_sec
                # Fall through to normal item processing below
            else:
                # No quantities → true section / chapter header
                try:
                    sec_id = odoo_call("project.main.item.line", "create",
                        [{"name": explicit_sec, "project_id": project_id}])
                    current_main_id = sec_id
                    sections += 1
                    last_name = explicit_sec
                except Exception as e:
                    errors.append(f"Row {i+1} section: {_err(e)}")
                continue

        # ── Carry-forward: merged-cell rows have qty/price but no name ──────────
        if not vals.get("name"):
            if last_name and (qty_pre or uc_pre or tc_pre):
                vals["name"] = last_name
                if last_item_code and not vals.get("item_code"):
                    vals["item_code"] = last_item_code

        if not vals.get("name"):
            skipped += 1
            continue

        # Update carry-forward state
        last_name = vals["name"]
        if vals.get("item_code"):
            last_item_code = vals["item_code"]

        qty        = _to_float(vals.get("quantity", 0))
        unit_cost  = _to_float(vals.get("unit_cost") or vals.get("unit_price", 0))
        total_cost = _to_float(vals.get("total_cost", 0))
        if qty and unit_cost and not total_cost:
            total_cost = qty * unit_cost
        elif qty and total_cost and not unit_cost:
            unit_cost = round(total_cost / qty, 4)

        # Resolve Many2one IDs provided directly in the file
        file_main_id  = _to_int_id(vals.get("main_item_line_id"))
        file_boq_id   = _to_int_id(vals.get("boq_item_id"))
        row_main_id   = file_main_id or current_main_id  # file ID wins, else auto-tracked

        # ── Auto-detect section header (no qty, short code without dash) ────
        if _is_section_header(vals, qty, unit_cost) and not file_main_id:
            code = str(vals.get("item_code", "")).strip()
            sec_name = f"[{code}] {vals['name']}" if code else vals["name"]
            try:
                sec_id = odoo_call("project.main.item.line", "create",
                    [{"name": sec_name, "project_id": project_id}])
                current_main_id = sec_id
                sections += 1
            except Exception as e:
                errors.append(f"Row {i+1} section: {_err(e)}")
            continue

        # Build display name
        item_name = vals["name"]
        if vals.get("item_code"):
            item_name = f"[{vals['item_code']}] {vals['name']}"

        # ── project.subcontracting.boq.line ──────────────────────────────────
        if "project.subcontracting.boq.line" in models_to_import:
            try:
                rec = {"name": item_name, "project_id": project_id}
                if boq_contract_id: rec["boq_contract_id"] = boq_contract_id
                if qty:             rec["quantity"]  = qty
                if unit_cost:       rec["unit_cost"] = unit_cost
                if vals.get("work_type"): rec["work_type"] = vals["work_type"]
                odoo_call("project.subcontracting.boq.line", "create", [rec])
                sub_ok += 1
            except Exception as e:
                errors.append(f"Row {i+1} subcontracting: {_err(e)}")
                logging.warning("subcontracting row %d: %s", i+1, e)

        # ── project.detailed.item.line ────────────────────────────────────────
        if "project.detailed.item.line" in models_to_import:
            try:
                rec2 = {"name": item_name, "project_id": project_id}
                if qty:          rec2["quantity"]  = qty
                if unit_cost:    rec2["unit_cost"] = unit_cost
                if row_main_id:  rec2["main_item_line_id"] = row_main_id
                if file_boq_id:  rec2["boq_item_id"] = file_boq_id
                odoo_call("project.detailed.item.line", "create", [rec2])
                det_ok += 1
            except Exception as e:
                err_msg = _err(e)
                # FK violation on boq_item_id → retry without it (file value not in DB)
                if file_boq_id and ("boq_item_id" in err_msg or "fkey" in err_msg.lower()):
                    try:
                        rec2.pop("boq_item_id", None)
                        odoo_call("project.detailed.item.line", "create", [rec2])
                        det_ok += 1
                        logging.info("Row %d: imported without boq_item_id (FK miss)", i+1)
                    except Exception as e2:
                        errors.append(f"Row {i+1} detailed: {_err(e2)}")
                        logging.warning("detailed row %d (retry): %s", i+1, e2)
                else:
                    errors.append(f"Row {i+1} detailed: {err_msg}")
                    logging.warning("detailed row %d: %s", i+1, e)

        if len(errors) >= 50:
            errors.append("...import stopped after 50 errors — check Railway logs")
            break

    imported = max(sub_ok, det_ok)

    return jsonify({
        "imported":         imported,
        "sub_ok":           sub_ok,
        "det_ok":           det_ok,
        "sections":         sections,
        "skipped":          skipped,
        "errors":           errors,
        "boq_contract_id":  boq_contract_id,
        "models_used":      list(models_to_import) + (
            ["project.main.item.line"] if sections > 0 else []),
    })


# ── WhatsApp (Twilio Sandbox) ──────────────────────────────────────────────────

def _twilio_send(to_number, body):
    """Send a WhatsApp message via Twilio REST API."""
    sid   = os.environ.get("TWILIO_ACCOUNT_SID", "")
    token = os.environ.get("TWILIO_AUTH_TOKEN", "")
    from_ = os.environ.get("TWILIO_WHATSAPP_FROM", "whatsapp:+14155238886")
    # Ensure whatsapp: prefix is present
    if from_ and not from_.startswith("whatsapp:"):
        from_ = "whatsapp:" + from_
    if not to_number.startswith("whatsapp:"):
        to_number = "whatsapp:" + to_number.lstrip("whatsapp:")
    if not sid or not token:
        logging.warning("Twilio credentials not set — cannot send WhatsApp reply")
        return
    if not _requests:
        logging.warning("requests library not installed")
        return
    auth = base64.b64encode(f"{sid}:{token}".encode()).decode()
    url  = f"https://api.twilio.com/2010-04-01/Accounts/{sid}/Messages.json"
    try:
        r = _requests.post(url,
            headers={"Authorization": f"Basic {auth}"},
            data={"From": from_, "To": to_number, "Body": body},
            timeout=15)
        if r.status_code >= 400:
            logging.error("Twilio API error %s: %s", r.status_code, r.text[:300])
    except Exception as e:
        logging.error("Twilio send error: %s", e)


def _md_to_whatsapp(text):
    """Convert markdown to WhatsApp-compatible formatting."""
    # **bold** → *bold*
    text = re.sub(r'\*\*(.+?)\*\*', r'*\1*', text)
    # # Heading → *Heading*
    text = re.sub(r'^#{1,4}\s+(.+)$', r'*\1*', text, flags=re.MULTILINE)
    # Remove code fences
    text = re.sub(r'```[^`]*```', '', text, flags=re.DOTALL)
    # Strip chart data lines
    text = re.sub(r'CHART_\w+:\{[^\n]+\}', '', text)
    # Collapse excess blank lines
    text = re.sub(r'\n{3,}', '\n\n', text)
    return text.strip()


def _split_whatsapp(text, max_len=4000):
    """Split message at paragraph boundaries so each chunk ≤ max_len chars."""
    if len(text) <= max_len:
        return [text]
    chunks = []
    while len(text) > max_len:
        pos = text.rfind('\n\n', 0, max_len)
        if pos == -1:
            pos = text.rfind('\n', 0, max_len)
        if pos == -1:
            pos = max_len
        chunks.append(text[:pos].strip())
        text = text[pos:].strip()
    if text:
        chunks.append(text)
    return chunks


def _run_ai_sync(user_text):
    """Run the full AI + Odoo tool-call loop synchronously. Returns answer text."""
    client, _model = _make_chat_client()
    if not client:
        return "❌ AI API key not configured on the server."

    messages = [
        {"role": "system", "content": get_system_prompt()},
        {"role": "user",   "content": user_text},
    ]

    for _ in range(3):
        _has_results = any(m.get("role") == "tool" for m in messages)
        call_kw = dict(
            model=_model, messages=messages, tools=TOOLS,
            tool_choice="none" if _has_results else "auto",
            max_tokens=1800, temperature=0.1,
        )
        try:
            resp = client.chat.completions.create(**call_kw)
        except Exception as e:
            return f"❌ AI error: {str(e)[:200]}"

        msg    = resp.choices[0].message
        finish = resp.choices[0].finish_reason

        asst: dict = {"role": "assistant", "content": msg.content or ""}
        if msg.tool_calls:
            asst["tool_calls"] = [
                {"id": tc.id, "type": "function",
                 "function": {"name": tc.function.name, "arguments": tc.function.arguments}}
                for tc in msg.tool_calls
            ]
        messages.append(asst)

        if finish in ("stop", "end_turn") or not msg.tool_calls:
            return msg.content or "لم أستطع الإجابة على هذا السؤال."

        for tc in msg.tool_calls:
            try:
                args = json.loads(tc.function.arguments)
            except Exception:
                args = {}
            result = run_tool(tc.function.name, args)
            messages.append({"role": "tool", "tool_call_id": tc.id, "content": result})

    return "لم أتمكن من إكمال الطلب. يرجى إعادة الصياغة."


def _transcribe_voice(media_url):
    """Download voice note from Twilio and transcribe with Groq Whisper."""
    groq_key = DEFAULT_GROQ_KEY
    if not groq_key or not _requests:
        return None

    sid   = os.environ.get("TWILIO_ACCOUNT_SID", "")
    token = os.environ.get("TWILIO_AUTH_TOKEN", "")

    # Download audio (Twilio requires auth to access media URLs)
    auth = base64.b64encode(f"{sid}:{token}".encode()).decode()
    resp = _requests.get(media_url,
                         headers={"Authorization": f"Basic {auth}"},
                         timeout=30)
    if resp.status_code != 200:
        logging.warning("Failed to download voice note: %s", resp.status_code)
        return None

    audio_bytes = resp.content
    content_type = resp.headers.get("Content-Type", "audio/ogg")
    # Derive a file extension from content type
    ext = "ogg"
    if "mpeg" in content_type or "mp3" in content_type:
        ext = "mp3"
    elif "mp4" in content_type or "m4a" in content_type:
        ext = "mp4"
    elif "wav" in content_type:
        ext = "wav"

    try:
        client_tmp = Groq(api_key=groq_key)
        transcript = client_tmp.audio.transcriptions.create(
            model="whisper-large-v3-turbo",
            file=(f"voice.{ext}", audio_bytes),
            response_format="text",
        )
        return str(transcript).strip()
    except Exception as e:
        logging.error("Whisper transcription failed: %s", e)
        return None


def _handle_whatsapp(from_number, user_text):
    """Background worker: run AI query then send WhatsApp reply."""
    logging.info("WhatsApp from %s: %s", from_number, user_text[:80])
    try:
        answer = _run_ai_sync(user_text)
        formatted = _md_to_whatsapp(answer)
        for chunk in _split_whatsapp(formatted):
            _twilio_send(from_number, chunk)
    except Exception as e:
        _twilio_send(from_number, f"❌ خطأ: {str(e)[:200]}")


@app.route("/whatsapp", methods=["POST"])
@app.route("/ai/whatsapp", methods=["POST"])
def whatsapp_webhook():
    """Twilio WhatsApp webhook — returns TwiML ack immediately, AI reply via REST API."""
    from_number  = request.form.get("From", "").strip()
    body         = request.form.get("Body", "").strip()
    num_media    = int(request.form.get("NumMedia", "0") or "0")
    media_url    = request.form.get("MediaUrl0", "").strip()
    media_type   = request.form.get("MediaContentType0", "").strip()

    logging.info("WA webhook: from=%s body=%r media=%s", from_number, body[:60], media_type)

    def _twiml(msg=None):
        if msg:
            safe = msg.replace('&','&amp;').replace('<','&lt;').replace('>','&gt;')
            xml = f'<?xml version="1.0" encoding="UTF-8"?><Response><Message>{safe}</Message></Response>'
        else:
            xml = '<?xml version="1.0" encoding="UTF-8"?><Response></Response>'
        return (xml, 200, {"Content-Type": "text/xml"})

    if not from_number:
        return _twiml()

    # ── Voice note ──────────────────────────────────────────────────────────
    if num_media > 0 and media_url and "audio" in media_type:
        def _handle_voice():
            try:
                transcript = _transcribe_voice(media_url)
                if not transcript:
                    _twilio_send(from_number, "❌ لم أتمكن من تفريغ الرسالة الصوتية. حاول مرة أخرى.")
                    return
                _twilio_send(from_number, f"🎙️ فهمت: {transcript}")
                answer = _run_ai_sync(transcript)
                for chunk in _split_whatsapp(_md_to_whatsapp(answer)):
                    _twilio_send(from_number, chunk)
            except Exception as e:
                logging.error("WA voice handler error: %s", e)
                _twilio_send(from_number, f"❌ خطأ: {str(e)[:200]}")
        threading.Thread(target=_handle_voice, daemon=True).start()
        # Immediate TwiML ack so Twilio doesn't retry
        return _twiml("🎤 جاري تفريغ الصوت وتحليل سؤالك…")

    # ── Text message ────────────────────────────────────────────────────────
    elif body:
        threading.Thread(
            target=_handle_whatsapp,
            args=(from_number, body),
            daemon=True
        ).start()
        # Immediate TwiML ack — user sees this in ~1 second, full answer follows
        return _twiml("⏳ جاري التحليل والاستعلام من Odoo…")

    return _twiml()


@app.route("/wa-test", methods=["GET"])
def wa_test():
    """Diagnostic: check Twilio credentials and optionally send a test message."""
    sid   = os.environ.get("TWILIO_ACCOUNT_SID", "")
    token = os.environ.get("TWILIO_AUTH_TOKEN", "")
    from_ = os.environ.get("TWILIO_WHATSAPP_FROM", "whatsapp:+14155238886")
    gemini = bool(GEMINI_API_KEY)
    groq  = bool(DEFAULT_GROQ_KEY)
    odoo  = bool(ODOO_API_KEY)
    to_   = request.args.get("to", "")   # ?to=whatsapp:+20XXXXXXXXX

    result = {
        "TWILIO_ACCOUNT_SID":   sid[:8] + "..." if sid else "❌ NOT SET",
        "TWILIO_AUTH_TOKEN":    "✅ set" if token else "❌ NOT SET",
        "TWILIO_WHATSAPP_FROM": from_ or "❌ NOT SET",
        "GEMINI_API_KEY":        "✅ set" if gemini else "❌ NOT SET",
        "GROQ_API_KEY (Whisper)": "✅ set" if groq else "⚠️ not set (voice only)",
        "ODOO_API_KEY":         "✅ set" if odoo else "❌ NOT SET",
        "webhook_url":          request.host_url.rstrip("/") + "/whatsapp",
    }

    if to_ and sid and token and _requests:
        if not from_.startswith("whatsapp:"):
            from_ = "whatsapp:" + from_
        if not to_.startswith("whatsapp:"):
            to_ = "whatsapp:" + to_
        auth = base64.b64encode(f"{sid}:{token}".encode()).decode()
        url  = f"https://api.twilio.com/2010-04-01/Accounts/{sid}/Messages.json"
        try:
            r = _requests.post(url,
                headers={"Authorization": f"Basic {auth}"},
                data={"From": from_, "To": to_, "Body": "🤖 Test from Odoo AI Assistant — connection OK!"},
                timeout=15)
            result["test_send"] = {"status": r.status_code, "body": r.text[:400]}
        except Exception as e:
            result["test_send"] = {"error": str(e)}
    elif to_:
        result["test_send"] = "❌ Cannot send — missing Twilio credentials or requests library"

    return jsonify(result)


@app.route("/whatsapp", methods=["GET"])
def whatsapp_info():
    return jsonify({"status": "ok", "webhook": "POST /whatsapp", "hint": "This endpoint accepts POST from Twilio only."})


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
